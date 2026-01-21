# -*- coding: utf-8 -*-
"""HWP 테이블 ↔ 엑셀 변환 모듈"""

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from dataclasses import dataclass
from typing import Dict, List, Optional, Any
from table.cell_position import CellPositionCalculator, CellPositionResult, CellRange

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
    from openpyxl.cell.cell import MergedCell
    from openpyxl.worksheet.page import PageMargins
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


@dataclass
class CellStyle:
    """셀 스타일 정보"""
    bg_color_rgb: tuple = None      # 배경색 (R, G, B)
    margin_left: int = 0            # 왼쪽 여백 (HWPUNIT)
    margin_right: int = 0           # 오른쪽 여백 (HWPUNIT)
    margin_top: int = 0             # 위쪽 여백 (HWPUNIT)
    margin_bottom: int = 0          # 아래쪽 여백 (HWPUNIT)
    # 글자 속성
    font_name: str = None           # 글꼴 이름
    font_size: float = 0            # 글자 크기 (pt)
    font_bold: bool = False         # 굵게
    font_italic: bool = False       # 기울임
    # 정렬
    align_horizontal: str = None    # 가로 정렬 (left, center, right)
    align_vertical: str = None      # 세로 정렬 (top, center, bottom)


@dataclass
class CellData:
    """셀 데이터"""
    list_id: int
    row: int
    col: int
    end_row: int
    end_col: int
    text: str
    rowspan: int = 1
    colspan: int = 1
    # 물리 좌표 (디버깅용)
    start_x: int = 0
    start_y: int = 0
    end_x: int = 0
    end_y: int = 0
    # 스타일 정보
    bg_color_rgb: tuple = None  # (R, G, B) 튜플
    style: CellStyle = None     # 상세 스타일


class TableExcelConverter:
    """HWP 테이블 ↔ 엑셀 변환기"""

    def __init__(self, hwp, debug: bool = False):
        self.hwp = hwp
        self.debug = debug
        self._calc = CellPositionCalculator(hwp, debug=debug)

    def _get_cell_text(self, list_id: int) -> str:
        """셀의 텍스트 내용 가져오기"""
        try:
            self.hwp.SetPos(list_id, 0, 0)

            # 셀 시작으로 이동
            self.hwp.MovePos(4, 0, 0)  # moveTopOfList

            # 셀 끝까지 선택
            self.hwp.HAction.Run("MoveSelBottomOfList")

            # 선택된 텍스트 가져오기
            text = self.hwp.GetTextFile("TEXT", "")

            # 선택 해제
            self.hwp.HAction.Run("Cancel")

            # 텍스트 정리
            if text:
                text = text.strip().replace('\r\n', ' ').replace('\r', ' ').replace('\n', ' ')
            return text or ""
        except Exception as e:
            if self.debug:
                print(f"[경고] list_id={list_id} 텍스트 추출 실패: {e}")
            return ""

    def validate_cell_positions(self, result: CellPositionResult = None) -> dict:
        """셀 위치 계산 결과 검증

        Returns:
            dict: {
                'valid': bool,
                'overlaps': [(cell1, cell2, overlap_coords), ...],
                'gaps': [(row, col), ...],  # 비어있는 좌표
                'total_cells': int,
                'coverage': float,  # 커버리지 비율
            }
        """
        if result is None:
            result = self._calc.calculate()

        overlaps = []
        coord_to_cell = {}  # (row, col) -> CellData

        # 모든 셀의 좌표 매핑 및 중복 검사
        for list_id, cell in result.cells.items():
            for r in range(cell.start_row, cell.end_row + 1):
                for c in range(cell.start_col, cell.end_col + 1):
                    coord = (r, c)
                    if coord in coord_to_cell:
                        existing = coord_to_cell[coord]
                        overlaps.append({
                            'coord': coord,
                            'cell1': {'list_id': existing.list_id, 'range': f"({existing.start_row},{existing.start_col})~({existing.end_row},{existing.end_col})"},
                            'cell2': {'list_id': list_id, 'range': f"({cell.start_row},{cell.start_col})~({cell.end_row},{cell.end_col})"},
                        })
                    else:
                        coord_to_cell[coord] = cell

        # 빈 좌표 확인
        gaps = []
        for r in range(result.max_row + 1):
            for c in range(result.max_col + 1):
                if (r, c) not in coord_to_cell:
                    gaps.append((r, c))

        total_coords = (result.max_row + 1) * (result.max_col + 1)
        covered_coords = len(coord_to_cell)
        coverage = covered_coords / total_coords if total_coords > 0 else 0

        validation = {
            'valid': len(overlaps) == 0 and len(gaps) == 0,
            'overlaps': overlaps,
            'gaps': gaps,
            'total_cells': len(result.cells),
            'total_coords': total_coords,
            'covered_coords': covered_coords,
            'coverage': coverage,
        }

        # 검증 결과 출력
        print(f"\n=== 셀 위치 검증 결과 ===")
        print(f"총 셀 수: {len(result.cells)}개")
        print(f"테이블 크기: {result.max_row + 1}행 x {result.max_col + 1}열 = {total_coords}좌표")
        print(f"커버리지: {covered_coords}/{total_coords} ({coverage*100:.1f}%)")

        if overlaps:
            print(f"\n[오류] 중복 좌표: {len(overlaps)}개")
            for ov in overlaps[:5]:
                print(f"  ({ov['coord'][0]},{ov['coord'][1]}): {ov['cell1']['range']} vs {ov['cell2']['range']}")
            if len(overlaps) > 5:
                print(f"  ... 외 {len(overlaps) - 5}개")

        if gaps:
            print(f"\n[경고] 빈 좌표: {len(gaps)}개")
            # 연속된 빈 영역 그룹화
            if len(gaps) <= 20:
                print(f"  {gaps}")
            else:
                print(f"  {gaps[:10]} ... 외 {len(gaps) - 10}개")

        if validation['valid']:
            print("\n검증 결과: 정상 [OK]")
        else:
            print("\n검증 결과: 문제 발견 [FAIL]")

        return validation

    def _get_cell_style(self, list_id: int) -> CellStyle:
        """셀 스타일 정보 추출"""
        style = CellStyle()

        try:
            self.hwp.SetPos(list_id, 0, 0)

            # 1. 배경색 및 여백 (CellBorderFill)
            pset = self.hwp.HParameterSet.HCellBorderFill
            self.hwp.HAction.GetDefault("CellBorderFill", pset.HSet)

            # 배경색
            bg_color = pset.FillAttr.WinBrushFaceColor
            if bg_color and bg_color > 0 and bg_color != 4294967295:
                b = (bg_color >> 16) & 0xFF
                g = (bg_color >> 8) & 0xFF
                r = bg_color & 0xFF
                style.bg_color_rgb = (r, g, b)

            # 셀 내부 여백 (FillAttr에서)
            fill_attr = pset.FillAttr
            if fill_attr:
                style.margin_left = getattr(fill_attr, 'InsideMarginLeft', 0) or 0
                style.margin_right = getattr(fill_attr, 'InsideMarginRight', 0) or 0
                style.margin_top = getattr(fill_attr, 'InsideMarginTop', 0) or 0
                style.margin_bottom = getattr(fill_attr, 'InsideMarginBottom', 0) or 0

            # 2. 글자 속성 (CharShape)
            try:
                char_shape = self.hwp.CharShape
                if char_shape:
                    # 글꼴 (한글 기준)
                    style.font_name = char_shape.Item("FaceNameHangul")
                    # 글자 크기 (HWPUNIT -> pt, 1pt = 100 HWPUNIT)
                    height = char_shape.Item("Height")
                    if height:
                        style.font_size = height / 100
                    # 굵게/기울임
                    style.font_bold = bool(char_shape.Item("Bold"))
                    style.font_italic = bool(char_shape.Item("Italic"))
            except:
                pass

            # 3. 정렬 (ParaShape)
            try:
                para_shape = self.hwp.ParaShape
                if para_shape:
                    align_val = para_shape.Item("Align")
                    # 0=양쪽, 1=왼쪽, 2=오른쪽, 3=가운데, 4=배분, 5=나눔
                    align_map = {0: 'justify', 1: 'left', 2: 'right', 3: 'center', 4: 'distribute', 5: 'divide'}
                    style.align_horizontal = align_map.get(align_val, 'left')
            except:
                pass

        except Exception as e:
            if self.debug:
                print(f"[경고] 셀 스타일 추출 실패 (list_id={list_id}): {e}")

        return style

    def _get_cell_bg_color(self, list_id: int) -> tuple:
        """셀 배경색 추출 (RGB 튜플 반환) - 호환성 유지"""
        style = self._get_cell_style(list_id)
        return style.bg_color_rgb

    def extract_table_data(self, max_cells: int = 1000, extract_style: bool = True) -> List[CellData]:
        """현재 테이블의 모든 셀 데이터 추출

        Args:
            max_cells: 최대 추출 셀 수
            extract_style: True면 배경색 등 스타일 정보도 추출
        """
        result = self._calc.calculate()

        cells_data = []
        count = 0
        for list_id, cell in result.cells.items():
            if count >= max_cells:
                print(f"[경고] 최대 셀 수({max_cells}) 도달, 중단")
                break

            text = self._get_cell_text(list_id)

            # 스타일 추출
            cell_style = None
            bg_color_rgb = None
            if extract_style:
                cell_style = self._get_cell_style(list_id)
                bg_color_rgb = cell_style.bg_color_rgb

            cells_data.append(CellData(
                list_id=list_id,
                row=cell.start_row,
                col=cell.start_col,
                end_row=cell.end_row,
                end_col=cell.end_col,
                text=text,
                rowspan=cell.rowspan,
                colspan=cell.colspan,
                start_x=cell.start_x,
                start_y=cell.start_y,
                end_x=cell.end_x,
                end_y=cell.end_y,
                bg_color_rgb=bg_color_rgb,
                style=cell_style,
            ))
            count += 1

            if self.debug and count % 50 == 0:
                print(f"  {count}개 셀 처리됨...")

        return cells_data

    def _get_hwp_page_margins(self):
        """한글 페이지 여백 정보 가져오기 (인치 단위로 변환)"""
        try:
            act = self.hwp.CreateAction("PageSetup")
            pset = act.CreateSet()
            act.GetDefault(pset)
            page_def = pset.Item("PageDef")

            # HWPUNIT -> 인치 변환 (7200 HWPUNIT = 1 inch)
            HWPUNIT_PER_INCH = 7200

            return {
                'left': page_def.Item("LeftMargin") / HWPUNIT_PER_INCH,
                'right': page_def.Item("RightMargin") / HWPUNIT_PER_INCH,
                'top': page_def.Item("TopMargin") / HWPUNIT_PER_INCH,
                'bottom': page_def.Item("BottomMargin") / HWPUNIT_PER_INCH,
                'header': page_def.Item("HeaderLen") / HWPUNIT_PER_INCH,
                'footer': page_def.Item("FooterLen") / HWPUNIT_PER_INCH,
                'orientation': 'landscape' if page_def.Item("Landscape") else 'portrait',
            }
        except Exception as e:
            if self.debug:
                print(f"[경고] 페이지 여백 추출 실패: {e}")
            return None

    def to_excel(self, filepath: str, sheet_name: str = "Sheet1", with_text: bool = True, show_cell_info: bool = False):
        """HWP 테이블을 엑셀 파일로 저장

        Args:
            filepath: 저장할 엑셀 파일 경로
            sheet_name: 시트 이름
            with_text: True면 셀 텍스트 포함, False면 셀 범위만 저장
            show_cell_info: True면 셀 속성 정보(list_id, 좌표, 병합정보) 표시
        """
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl이 설치되어 있지 않습니다. pip install openpyxl")

        # 셀 위치 계산
        result = self._calc.calculate()

        # 텍스트 포함 여부에 따라 데이터 준비
        if with_text:
            cells_data = self.extract_table_data()
        else:
            # 텍스트 없이 셀 범위만
            cells_data = [
                CellData(
                    list_id=cell.list_id,
                    row=cell.start_row,
                    col=cell.start_col,
                    end_row=cell.end_row,
                    end_col=cell.end_col,
                    text="",
                    rowspan=cell.rowspan,
                    colspan=cell.colspan,
                    start_x=cell.start_x,
                    start_y=cell.start_y,
                    end_x=cell.end_x,
                    end_y=cell.end_y,
                )
                for cell in result.cells.values()
            ]

        # 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # 한글 페이지 여백 가져오기
        hwp_margins = self._get_hwp_page_margins()

        # 인쇄 설정: 페이지에 맞춤
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1  # 1페이지 너비에 맞춤
        ws.page_setup.fitToHeight = 0  # 높이는 자동

        if hwp_margins:
            # 한글 페이지 방향 적용
            ws.page_setup.orientation = hwp_margins['orientation']

            # 한글 여백을 엑셀에 적용 (인치 단위)
            ws.page_margins = PageMargins(
                left=hwp_margins['left'],
                right=hwp_margins['right'],
                top=hwp_margins['top'],
                bottom=hwp_margins['bottom'],
                header=hwp_margins['header'],
                footer=hwp_margins['footer']
            )

            if self.debug:
                print(f"페이지 설정: {hwp_margins['orientation']}, "
                      f"여백(인치) L={hwp_margins['left']:.2f}, R={hwp_margins['right']:.2f}, "
                      f"T={hwp_margins['top']:.2f}, B={hwp_margins['bottom']:.2f}")
        else:
            # 기본값 사용
            ws.page_setup.orientation = 'portrait'
            ws.page_margins = PageMargins(
                left=0.5, right=0.5,
                top=0.5, bottom=0.5,
                header=0.3, footer=0.3
            )

        # 테두리 스타일
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 병합 영역 검증 및 적용
        merged_coords = set()  # 이미 병합된 좌표 추적
        merge_conflicts = []   # 충돌 목록
        merge_success = []     # 성공 목록

        # 셀을 (row, col) 순서로 정렬하여 병합
        sorted_cells = sorted(cells_data, key=lambda c: (c.row, c.col))

        for cell_data in sorted_cells:
            if cell_data.rowspan > 1 or cell_data.colspan > 1:
                excel_row = cell_data.row + 1
                excel_col = cell_data.col + 1
                end_row = cell_data.end_row + 1
                end_col = cell_data.end_col + 1

                # 병합 영역이 이미 사용 중인지 확인
                conflict_coords = []
                for r in range(cell_data.row, cell_data.end_row + 1):
                    for c in range(cell_data.col, cell_data.end_col + 1):
                        if (r, c) in merged_coords:
                            conflict_coords.append((r, c))

                if conflict_coords:
                    merge_conflicts.append({
                        'cell': cell_data,
                        'conflicts': conflict_coords,
                    })
                    if self.debug:
                        print(f"[충돌] list_id={cell_data.list_id} ({cell_data.row},{cell_data.col})~({cell_data.end_row},{cell_data.end_col}): "
                              f"이미 병합된 좌표 {conflict_coords[:3]}{'...' if len(conflict_coords) > 3 else ''}")
                else:
                    # 병합 적용
                    ws.merge_cells(
                        start_row=excel_row,
                        start_column=excel_col,
                        end_row=end_row,
                        end_column=end_col
                    )
                    # 병합된 좌표 기록
                    for r in range(cell_data.row, cell_data.end_row + 1):
                        for c in range(cell_data.col, cell_data.end_col + 1):
                            merged_coords.add((r, c))
                    merge_success.append(cell_data)

        # 병합 검증 결과 출력
        if merge_conflicts:
            print(f"\n=== 병합 충돌 감지: {len(merge_conflicts)}개 ===")
            for conflict in merge_conflicts[:10]:  # 최대 10개만 출력
                cell = conflict['cell']
                print(f"  list_id={cell.list_id}: ({cell.row},{cell.col})~({cell.end_row},{cell.end_col}) "
                      f"- 충돌 좌표: {conflict['conflicts'][:5]}")
            if len(merge_conflicts) > 10:
                print(f"  ... 외 {len(merge_conflicts) - 10}개")

        if self.debug:
            print(f"\n병합 성공: {len(merge_success)}개, 충돌: {len(merge_conflicts)}개")

        # 셀 데이터 쓰기 (병합 후 좌상단 셀에만 값 입력)
        skipped_cells = []
        for cell_data in cells_data:
            excel_row = cell_data.row + 1
            excel_col = cell_data.col + 1

            cell = ws.cell(row=excel_row, column=excel_col)

            # MergedCell인 경우 건너뛰기 (이미 다른 셀에 병합됨)
            if isinstance(cell, MergedCell):
                skipped_cells.append(cell_data)
                continue

            # 셀 내용 결정
            if show_cell_info:
                # 셀 속성 정보 표시
                info_lines = [f"id={cell_data.list_id}"]
                if cell_data.rowspan > 1 or cell_data.colspan > 1:
                    info_lines.append(f"({cell_data.row},{cell_data.col})~({cell_data.end_row},{cell_data.end_col})")
                    info_lines.append(f"span={cell_data.rowspan}x{cell_data.colspan}")
                else:
                    info_lines.append(f"({cell_data.row},{cell_data.col})")
                cell.value = "\n".join(info_lines)
            else:
                cell.value = cell_data.text

            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='center')

            # 배경색 적용
            if cell_data.bg_color_rgb:
                r, g, b = cell_data.bg_color_rgb
                hex_color = f"{r:02X}{g:02X}{b:02X}"
                cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

            # 병합 영역 테두리 및 배경색 적용
            if cell_data.rowspan > 1 or cell_data.colspan > 1:
                end_row = cell_data.end_row + 1
                end_col = cell_data.end_col + 1
                for r in range(excel_row, end_row + 1):
                    for c in range(excel_col, end_col + 1):
                        try:
                            merged_cell = ws.cell(row=r, column=c)
                            merged_cell.border = thin_border
                            # 병합 영역에도 배경색 적용
                            if cell_data.bg_color_rgb:
                                r_val, g_val, b_val = cell_data.bg_color_rgb
                                hex_color = f"{r_val:02X}{g_val:02X}{b_val:02X}"
                                merged_cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                        except AttributeError:
                            pass  # MergedCell은 건너뜀

        # 스킵된 셀 출력
        if skipped_cells:
            print(f"\n=== 스킵된 셀 (충돌): {len(skipped_cells)}개 ===")
            for cell in skipped_cells[:10]:
                print(f"  list_id={cell.list_id}: ({cell.row},{cell.col})~({cell.end_row},{cell.end_col})")
            if len(skipped_cells) > 10:
                print(f"  ... 외 {len(skipped_cells) - 10}개")

        # 열 너비 조정 (HWP 실제 너비 기준)
        # 엑셀 열 너비는 "문자 수" 단위 (기본 폰트 기준 약 7pt = 1문자)
        # 1 pt = 100 HWPUNIT, 1 문자 ≈ 7 pt = 700 HWPUNIT
        x_levels = result.x_levels
        if len(x_levels) > 1:
            HWPUNIT_PER_CHAR = 700  # 약 7pt per character
            for col_idx in range(len(x_levels) - 1):
                col_width_hwp = x_levels[col_idx + 1] - x_levels[col_idx]
                col_width_chars = col_width_hwp / HWPUNIT_PER_CHAR
                column_letter = get_column_letter(col_idx + 1)
                ws.column_dimensions[column_letter].width = max(col_width_chars, 2)

        # 행 높이 조정 (HWP 실제 높이 기준)
        # 엑셀 행 높이는 포인트(pt) 단위
        # 1 pt = 100 HWPUNIT
        y_levels = result.y_levels
        if len(y_levels) > 1:
            HWPUNIT_PER_PT = 100
            for row_idx in range(len(y_levels) - 1):
                row_height_hwp = y_levels[row_idx + 1] - y_levels[row_idx]
                row_height_pt = row_height_hwp / HWPUNIT_PER_PT
                ws.row_dimensions[row_idx + 1].height = max(row_height_pt, 12)

        if self.debug:
            print(f"\n크기 설정: {len(x_levels)-1}열, {len(y_levels)-1}행")

        # _meta 시트 생성 (셀 속성 정보)
        ws_meta = wb.create_sheet(title="_meta")

        # 헤더 (위치 + 스타일 정보)
        headers = [
            "list_id", "row", "col", "end_row", "end_col", "rowspan", "colspan",
            "start_x", "start_y", "end_x", "end_y", "is_merged", "status",
            # 스타일 정보
            "bg_color", "margin_L", "margin_R", "margin_T", "margin_B",
            "font_name", "font_size", "font_bold", "font_italic", "align_h"
        ]
        for col_idx, header in enumerate(headers, 1):
            ws_meta.cell(row=1, column=col_idx, value=header)
            ws_meta.cell(row=1, column=col_idx).font = Font(bold=True)

        # 셀 데이터 작성
        sorted_cells = sorted(cells_data, key=lambda c: (c.row, c.col))
        for row_idx, cell_data in enumerate(sorted_cells, 2):
            # 상태 결정
            if cell_data in skipped_cells:
                status = "SKIPPED"
            elif any(c['cell'] == cell_data for c in merge_conflicts):
                status = "CONFLICT"
            else:
                status = "OK"

            # 위치 정보
            ws_meta.cell(row=row_idx, column=1, value=cell_data.list_id)
            ws_meta.cell(row=row_idx, column=2, value=cell_data.row)
            ws_meta.cell(row=row_idx, column=3, value=cell_data.col)
            ws_meta.cell(row=row_idx, column=4, value=cell_data.end_row)
            ws_meta.cell(row=row_idx, column=5, value=cell_data.end_col)
            ws_meta.cell(row=row_idx, column=6, value=cell_data.rowspan)
            ws_meta.cell(row=row_idx, column=7, value=cell_data.colspan)
            ws_meta.cell(row=row_idx, column=8, value=cell_data.start_x)
            ws_meta.cell(row=row_idx, column=9, value=cell_data.start_y)
            ws_meta.cell(row=row_idx, column=10, value=cell_data.end_x)
            ws_meta.cell(row=row_idx, column=11, value=cell_data.end_y)
            ws_meta.cell(row=row_idx, column=12, value="Y" if cell_data.rowspan > 1 or cell_data.colspan > 1 else "N")
            ws_meta.cell(row=row_idx, column=13, value=status)

            # 스타일 정보
            style = cell_data.style
            if style:
                # 배경색 (RGB hex)
                if style.bg_color_rgb:
                    r, g, b = style.bg_color_rgb
                    ws_meta.cell(row=row_idx, column=14, value=f"#{r:02X}{g:02X}{b:02X}")
                # 여백 (HWPUNIT)
                ws_meta.cell(row=row_idx, column=15, value=style.margin_left)
                ws_meta.cell(row=row_idx, column=16, value=style.margin_right)
                ws_meta.cell(row=row_idx, column=17, value=style.margin_top)
                ws_meta.cell(row=row_idx, column=18, value=style.margin_bottom)
                # 글꼴
                ws_meta.cell(row=row_idx, column=19, value=style.font_name)
                ws_meta.cell(row=row_idx, column=20, value=style.font_size)
                ws_meta.cell(row=row_idx, column=21, value="Y" if style.font_bold else "")
                ws_meta.cell(row=row_idx, column=22, value="Y" if style.font_italic else "")
                ws_meta.cell(row=row_idx, column=23, value=style.align_horizontal)

        # _meta 열 너비 조정
        col_widths = [8, 5, 5, 7, 7, 7, 7, 8, 8, 8, 8, 9, 9,
                      10, 8, 8, 8, 8, 12, 8, 8, 8, 8]
        for col_idx, width in enumerate(col_widths, 1):
            ws_meta.column_dimensions[get_column_letter(col_idx)].width = width

        # _meta 시트 인쇄 설정 (세로, 1페이지 너비 맞춤)
        ws_meta.page_setup.orientation = 'portrait'
        ws_meta.page_setup.fitToPage = True
        ws_meta.page_setup.fitToWidth = 1
        ws_meta.page_setup.fitToHeight = 0
        ws_meta.page_margins = PageMargins(
            left=0.5, right=0.5, top=0.5, bottom=0.5,
            header=0.3, footer=0.3
        )

        # _page 시트 생성 (페이지 설정 정보)
        ws_page = wb.create_sheet(title="_page")

        # 페이지 정보 헤더 스타일
        header_font = Font(bold=True)

        # 한글 페이지 설정 정보 수집
        page_info = []

        # 단위 변환 상수
        HWPUNIT_PER_INCH = 7200
        HWPUNIT_PER_CM = 7200 / 2.54

        try:
            act = self.hwp.CreateAction("PageSetup")
            pset = act.CreateSet()
            act.GetDefault(pset)
            page_def = pset.Item("PageDef")

            # 용지 크기
            paper_width = page_def.Item("PaperWidth")
            paper_height = page_def.Item("PaperHeight")
            landscape = page_def.Item("Landscape")

            # 여백
            left_margin = page_def.Item("LeftMargin")
            right_margin = page_def.Item("RightMargin")
            top_margin = page_def.Item("TopMargin")
            bottom_margin = page_def.Item("BottomMargin")
            header_len = page_def.Item("HeaderLen")
            footer_len = page_def.Item("FooterLen")
            gutter_len = page_def.Item("GutterLen")

            page_info = [
                ("항목", "값 (HWPUNIT)", "값 (cm)", "값 (inch)"),
                ("용지 너비", paper_width, f"{paper_width/HWPUNIT_PER_CM:.2f}", f"{paper_width/HWPUNIT_PER_INCH:.2f}"),
                ("용지 높이", paper_height, f"{paper_height/HWPUNIT_PER_CM:.2f}", f"{paper_height/HWPUNIT_PER_INCH:.2f}"),
                ("용지 방향", "가로" if landscape else "세로", "", ""),
                ("왼쪽 여백", left_margin, f"{left_margin/HWPUNIT_PER_CM:.2f}", f"{left_margin/HWPUNIT_PER_INCH:.2f}"),
                ("오른쪽 여백", right_margin, f"{right_margin/HWPUNIT_PER_CM:.2f}", f"{right_margin/HWPUNIT_PER_INCH:.2f}"),
                ("위쪽 여백", top_margin, f"{top_margin/HWPUNIT_PER_CM:.2f}", f"{top_margin/HWPUNIT_PER_INCH:.2f}"),
                ("아래쪽 여백", bottom_margin, f"{bottom_margin/HWPUNIT_PER_CM:.2f}", f"{bottom_margin/HWPUNIT_PER_INCH:.2f}"),
                ("머리말", header_len, f"{header_len/HWPUNIT_PER_CM:.2f}", f"{header_len/HWPUNIT_PER_INCH:.2f}"),
                ("꼬리말", footer_len, f"{footer_len/HWPUNIT_PER_CM:.2f}", f"{footer_len/HWPUNIT_PER_INCH:.2f}"),
                ("제본 여백", gutter_len, f"{gutter_len/HWPUNIT_PER_CM:.2f}", f"{gutter_len/HWPUNIT_PER_INCH:.2f}"),
                ("", "", "", ""),
                ("본문 너비", paper_width - left_margin - right_margin - gutter_len,
                 f"{(paper_width - left_margin - right_margin - gutter_len)/HWPUNIT_PER_CM:.2f}",
                 f"{(paper_width - left_margin - right_margin - gutter_len)/HWPUNIT_PER_INCH:.2f}"),
                ("본문 높이", paper_height - top_margin - bottom_margin - header_len - footer_len,
                 f"{(paper_height - top_margin - bottom_margin - header_len - footer_len)/HWPUNIT_PER_CM:.2f}",
                 f"{(paper_height - top_margin - bottom_margin - header_len - footer_len)/HWPUNIT_PER_INCH:.2f}"),
                ("", "", "", ""),
                ("테이블 행 수", len(y_levels) - 1, "", ""),
                ("테이블 열 수", len(x_levels) - 1, "", ""),
                ("테이블 셀 수", len(cells_data), "", ""),
            ]
        except Exception as e:
            page_info = [
                ("항목", "값", "", ""),
                ("오류", str(e), "", ""),
            ]

        # 페이지 정보 기록
        for row_idx, row_data in enumerate(page_info, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_page.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.font = header_font

        # _page 열 너비 조정
        ws_page.column_dimensions['A'].width = 15
        ws_page.column_dimensions['B'].width = 15
        ws_page.column_dimensions['C'].width = 12
        ws_page.column_dimensions['D'].width = 12

        # _page 시트 인쇄 설정 (세로, 1페이지에 맞춤)
        ws_page.page_setup.orientation = 'portrait'
        ws_page.page_setup.fitToPage = True
        ws_page.page_setup.fitToWidth = 1
        ws_page.page_setup.fitToHeight = 1
        ws_page.page_margins = PageMargins(
            left=0.5, right=0.5, top=0.5, bottom=0.5,
            header=0.3, footer=0.3
        )

        wb.save(filepath)
        print(f"\n엑셀 파일 저장: {filepath}")
        print(f"  - 메인 시트: {sheet_name}")
        print(f"  - 메타 시트: _meta ({len(cells_data)}개 셀 정보)")
        print(f"  - 페이지 시트: _page (페이지/여백 설정)")
        return filepath

    def to_dict(self) -> Dict[tuple, str]:
        """테이블을 딕셔너리로 변환 {(row, col): text}"""
        cells_data = self.extract_table_data()
        result = {}

        for cell_data in cells_data:
            # 병합 셀은 모든 좌표에 같은 텍스트
            for r in range(cell_data.row, cell_data.end_row + 1):
                for c in range(cell_data.col, cell_data.end_col + 1):
                    result[(r, c)] = cell_data.text

        return result

    def to_2d_array(self) -> List[List[str]]:
        """테이블을 2차원 배열로 변환"""
        result = self._calc.calculate()
        cells_data = self.extract_table_data()

        # 빈 2D 배열 생성
        rows = result.max_row + 1
        cols = result.max_col + 1
        array = [["" for _ in range(cols)] for _ in range(rows)]

        # 데이터 채우기
        for cell_data in cells_data:
            for r in range(cell_data.row, cell_data.end_row + 1):
                for c in range(cell_data.col, cell_data.end_col + 1):
                    array[r][c] = cell_data.text

        return array

    def print_table(self):
        """테이블 내용을 콘솔에 출력"""
        array = self.to_2d_array()
        result = self._calc.calculate()

        print(f"\n=== 테이블 내용 ({result.max_row + 1}행 x {result.max_col + 1}열) ===")
        for row_idx, row in enumerate(array):
            row_str = " | ".join(cell[:10].ljust(10) if cell else "".ljust(10) for cell in row)
            print(f"[{row_idx:2d}] {row_str}")


# 직접 실행 시
if __name__ == "__main__":
    from cursor import get_hwp_instance

    hwp = get_hwp_instance()
    if not hwp:
        print("[오류] 한글이 실행 중이지 않습니다.")
        exit(1)

    converter = TableExcelConverter(hwp, debug=True)

    try:
        # 1. 셀 위치 계산
        result = converter._calc.calculate()
        converter._calc.print_summary(result)

        # 2. 셀 위치 검증
        validation = converter.validate_cell_positions(result)

        # 3. 엑셀로 저장 (셀 속성 정보 포함)
        if HAS_OPENPYXL:
            import datetime
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            converter.to_excel(f"C:\\win32hwp\\output_table_{timestamp}.xlsx", with_text=False, show_cell_info=True)
        else:
            print("[경고] openpyxl이 없어서 엑셀 저장을 건너뜁니다.")
            print("설치: pip install openpyxl")

    except ValueError as e:
        print(f"[오류] {e}")
