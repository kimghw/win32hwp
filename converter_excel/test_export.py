# -*- coding: utf-8 -*-
"""통합 테스트: 페이지 정보 + 셀 스타일 → 동일 엑셀 파일 저장

실행: cmd.exe /c "cd /d C:\\win32hwp && python converter_excel\\test_export.py"
"""

import sys
import os
import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from cursor import get_hwp_instance
from table.table_info import TableInfo, MOVE_RIGHT_OF_CELL, MOVE_DOWN_OF_CELL
from table.cell_position import CellPositionCalculator

try:
    from openpyxl import Workbook
    from openpyxl.worksheet.page import PageMargins
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("[오류] openpyxl이 필요합니다: pip install openpyxl")
    exit(1)

from converter_excel.match_page import (
    extract_page_info,
    write_page_info_to_sheet,
    apply_page_margins_to_excel,
)
from converter_excel.match_cell import (
    extract_cell_style,
    get_cell_text,
    write_cell_styles_to_sheet,
    write_row_col_sizes_to_sheet,
    apply_cell_style_to_excel_cell,
    CellStyleData,
)


def main():
    # 출력 파일 경로 (고정)
    OUTPUT_FILE = "C:\\win32hwp\\export_test.xlsx"

    hwp = get_hwp_instance()
    if not hwp:
        print("[오류] 한글이 실행 중이지 않습니다.")
        return

    table_info = TableInfo(hwp, debug=False)
    if not table_info.is_in_table():
        print("[오류] 커서가 표 안에 있지 않습니다.")
        return

    print("=" * 60)
    print("통합 테스트: 페이지 + 셀 스타일 → 엑셀")
    print("=" * 60)

    # ============================================================
    # 1. 페이지 정보 추출
    # ============================================================
    print("\n[1] 페이지 정보 추출...")
    page_result = extract_page_info(hwp)

    if page_result.success:
        meta = page_result.page_meta
        print(f"    용지: {meta.page_size.width/2834.6:.1f} x {meta.page_size.height/2834.6:.1f} cm ({meta.page_size.orientation})")
        print(f"    여백: 좌={meta.margin.left/2834.6:.1f}cm, 우={meta.margin.right/2834.6:.1f}cm, "
              f"상={meta.margin.top/2834.6:.1f}cm, 하={meta.margin.bottom/2834.6:.1f}cm")
    else:
        print(f"    [오류] {page_result.error}")
        return

    # ============================================================
    # 2. 셀 위치 계산 (CellPositionCalculator 사용)
    # ============================================================
    print("\n[2] 셀 위치 계산...")
    calc = CellPositionCalculator(hwp, debug=False)
    pos_result = calc.calculate()

    print(f"    행 수: {len(pos_result.y_levels) - 1}")
    print(f"    열 수: {len(pos_result.x_levels) - 1}")
    print(f"    셀 수: {len(pos_result.cells)}")

    # ============================================================
    # 3. 셀 스타일 추출
    # ============================================================
    print("\n[3] 셀 스타일 추출...")
    cells_data = []
    colored_count = 0

    for list_id, cell_range in pos_result.cells.items():
        style = extract_cell_style(hwp, list_id)
        style.list_id = list_id
        style.row = cell_range.start_row
        style.col = cell_range.start_col
        style.end_row = cell_range.end_row
        style.end_col = cell_range.end_col
        style.rowspan = cell_range.rowspan
        style.colspan = cell_range.colspan
        style.x = cell_range.start_x
        style.y = cell_range.start_y
        style.width = cell_range.end_x - cell_range.start_x
        style.height = cell_range.end_y - cell_range.start_y
        style.text = get_cell_text(hwp, list_id)

        if style.bg_color_rgb:
            colored_count += 1

        cells_data.append(style)

    print(f"    배경색 있는 셀: {colored_count}개")

    # 행 높이 / 열 너비 계산
    y_levels = pos_result.y_levels
    x_levels = pos_result.x_levels
    row_heights = [y_levels[i+1] - y_levels[i] for i in range(len(y_levels) - 1)]
    col_widths = [x_levels[i+1] - x_levels[i] for i in range(len(x_levels) - 1)]

    # ============================================================
    # 4. 엑셀 워크북 생성
    # ============================================================
    print("\n[4] 엑셀 파일 생성...")
    wb = Workbook()

    # ------------------------------------------------------------
    # Sheet 1: 표 데이터 (메인)
    # ------------------------------------------------------------
    ws_main = wb.active
    ws_main.title = "표"

    # 한글 페이지 설정 적용
    apply_page_margins_to_excel(ws_main, page_result.page_meta)
    ws_main.page_setup.fitToPage = True
    ws_main.page_setup.fitToWidth = 1
    ws_main.page_setup.fitToHeight = 0

    # 테두리 스타일
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 병합 트래킹
    merged_cells = set()

    # 셀 데이터 기록
    for cell_data in cells_data:
        excel_row = cell_data.row + 1
        excel_col = cell_data.col + 1

        # 이미 병합된 영역 스킵
        if (excel_row, excel_col) in merged_cells:
            continue

        cell = ws_main.cell(row=excel_row, column=excel_col)
        cell.value = cell_data.text
        cell.border = thin_border

        # 스타일 적용
        apply_cell_style_to_excel_cell(ws_main, excel_row, excel_col, cell_data)

        # 병합 처리
        if cell_data.rowspan > 1 or cell_data.colspan > 1:
            end_row = cell_data.end_row + 1
            end_col = cell_data.end_col + 1

            ws_main.merge_cells(
                start_row=excel_row,
                start_column=excel_col,
                end_row=end_row,
                end_column=end_col
            )

            # 병합 영역 트래킹 및 테두리/배경 적용
            for r in range(excel_row, end_row + 1):
                for c in range(excel_col, end_col + 1):
                    merged_cells.add((r, c))
                    try:
                        merged_cell = ws_main.cell(row=r, column=c)
                        merged_cell.border = thin_border
                        if cell_data.bg_color_rgb:
                            rgb = cell_data.bg_color_rgb
                            hex_color = f"{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}"
                            merged_cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                    except:
                        pass

    # 열 너비 설정 (HWPUNIT -> 문자)
    for col_idx, width in enumerate(col_widths):
        col_letter = get_column_letter(col_idx + 1)
        ws_main.column_dimensions[col_letter].width = max(width / 700, 2)

    # 행 높이 설정 (HWPUNIT -> pt)
    for row_idx, height in enumerate(row_heights):
        ws_main.row_dimensions[row_idx + 1].height = max(height / 100, 12)

    print(f"    [표] 시트 생성 완료")

    # ------------------------------------------------------------
    # Sheet 2: 페이지 설정
    # ------------------------------------------------------------
    ws_page = wb.create_sheet(title="_page")
    write_page_info_to_sheet(ws_page, page_result.page_meta)
    print(f"    [_page] 시트 생성 완료")

    # ------------------------------------------------------------
    # Sheet 3: 셀 스타일 정보
    # ------------------------------------------------------------
    ws_cells = wb.create_sheet(title="_cells")
    write_cell_styles_to_sheet(ws_cells, cells_data, row_heights, col_widths)
    print(f"    [_cells] 시트 생성 완료")

    # ------------------------------------------------------------
    # Sheet 4: 행/열 크기
    # ------------------------------------------------------------
    ws_sizes = wb.create_sheet(title="_sizes")
    write_row_col_sizes_to_sheet(ws_sizes, row_heights, col_widths)
    print(f"    [_sizes] 시트 생성 완료")

    # ============================================================
    # 5. 저장
    # ============================================================
    wb.save(OUTPUT_FILE)

    print("\n" + "=" * 60)
    print(f"저장 완료: {OUTPUT_FILE}")
    print("=" * 60)
    print(f"\n시트 구성:")
    print(f"  - 표       : 한글 표 데이터 (배경색, 병합 포함)")
    print(f"  - _page    : 페이지 설정 (용지, 여백)")
    print(f"  - _cells   : 셀 스타일 정보 (배경색, 글꼴, 정렬, 테두리)")
    print(f"  - _sizes   : 행 높이 / 열 너비")


if __name__ == "__main__":
    main()
