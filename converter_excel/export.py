# -*- coding: utf-8 -*-
"""HWP 표 → Excel 변환 메인 실행 파일

사용법:
    cmd.exe /c "cd /d C:\\win32hwp && python converter_excel\\export.py"
    cmd.exe /c "cd /d C:\\win32hwp && python converter_excel\\export.py --config my_config.yaml"
    cmd.exe /c "cd /d C:\\win32hwp && python converter_excel\\export.py --output result.xlsx"
"""

import sys
import os
import argparse
import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from cursor import get_hwp_instance
from table.table_info import TableInfo
from table.cell_position import CellPositionCalculator

try:
    from openpyxl import Workbook
    from openpyxl.worksheet.page import PageMargins
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Protection
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.protection import SheetProtection
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from converter_excel.config import load_config, get_default_config, ExportConfig
from converter_excel.match_page import extract_page_info, write_page_info_to_sheet
from converter_excel.match_cell import (
    extract_cell_style,
    get_cell_text,
    write_cell_styles_to_sheet,
    write_row_col_sizes_to_sheet,
    CellStyleData,
)
from converter_excel.field_name import generate_field_names, set_cell_field_names


class HwpToExcelExporter:
    """HWP 표 → Excel 변환기"""

    def __init__(self, config: ExportConfig = None):
        self.config = config or get_default_config()
        self.hwp = None
        self.table_info = None

    def connect(self) -> bool:
        """한글 인스턴스 연결"""
        self.hwp = get_hwp_instance()
        if not self.hwp:
            print("[오류] 한글이 실행 중이지 않습니다.")
            return False

        self.table_info = TableInfo(self.hwp, debug=False)
        if not self.table_info.is_in_table():
            print("[오류] 커서가 표 안에 있지 않습니다.")
            return False

        return True

    def export(self, output_path: str, sheet_name: str = "표") -> bool:
        """변환 실행

        Args:
            output_path: 출력 파일 경로
            sheet_name: 메인 시트 이름

        Returns:
            성공 여부
        """
        if not HAS_OPENPYXL:
            print("[오류] openpyxl이 필요합니다: pip install openpyxl")
            return False

        if not self.hwp:
            if not self.connect():
                return False

        cfg = self.config
        print("=" * 60)
        print("HWP 표 → Excel 변환")
        print("=" * 60)

        # ---------------------------------------------------------------------
        # 1. 페이지 정보 추출
        # ---------------------------------------------------------------------
        print("\n[1] 페이지 정보 추출...")
        page_result = extract_page_info(self.hwp)

        if page_result.success:
            meta = page_result.page_meta
            print(f"    용지: {meta.page_size.width/2834.6:.1f} x {meta.page_size.height/2834.6:.1f} cm ({meta.page_size.orientation})")
        else:
            print(f"    [경고] {page_result.error}")

        # ---------------------------------------------------------------------
        # 2. 셀 위치 계산
        # ---------------------------------------------------------------------
        print("\n[2] 셀 위치 계산...")
        calc = CellPositionCalculator(self.hwp, debug=False)
        pos_result = calc.calculate()

        print(f"    행 수: {len(pos_result.y_levels) - 1}")
        print(f"    열 수: {len(pos_result.x_levels) - 1}")
        print(f"    셀 수: {len(pos_result.cells)}")

        # ---------------------------------------------------------------------
        # 3. 셀 스타일 추출
        # ---------------------------------------------------------------------
        print("\n[3] 셀 스타일 추출...")
        cells_data = []
        colored_count = 0

        for list_id, cell_range in pos_result.cells.items():
            style = extract_cell_style(self.hwp, list_id)
            style.list_id = list_id
            style.row = cell_range.start_row
            style.col = cell_range.start_col
            style.end_row = cell_range.end_row
            style.end_col = cell_range.end_col
            style.rowspan = cell_range.rowspan
            style.colspan = cell_range.colspan
            style.x = cell_range.start_x
            style.y = cell_range.start_y
            style.width = cell_range.end_x - cell_range.start_x
            style.height = cell_range.end_y - cell_range.start_y
            style.text = get_cell_text(self.hwp, list_id)

            if style.bg_color_rgb:
                colored_count += 1

            cells_data.append(style)

        print(f"    배경색 있는 셀: {colored_count}개")

        # 행/열 크기
        y_levels = pos_result.y_levels
        x_levels = pos_result.x_levels
        row_heights = [y_levels[i+1] - y_levels[i] for i in range(len(y_levels) - 1)]
        col_widths = [x_levels[i+1] - x_levels[i] for i in range(len(x_levels) - 1)]

        # ---------------------------------------------------------------------
        # 4. 필드 생성 (설정에 따라)
        # ---------------------------------------------------------------------
        if cfg.field.enabled:
            print("\n[4] 필드 이름 생성...")
            fields = generate_field_names(
                self.hwp, cells_data,
                tolerance=cfg.field.naming.pattern.tolerance
            )
            print(f"    필드 수: {len(fields)}개")

            # 한글 셀에 필드 설정
            if cfg.field.set_hwp_field:
                success_count = set_cell_field_names(self.hwp, fields)
                print(f"    한글 셀 필드 설정: {success_count}/{len(fields)}개")

            # cells_data에 필드 정보 추가
            field_map = {f.list_id: f for f in fields}
            for cell_data in cells_data:
                if cell_data.list_id in field_map:
                    field_info = field_map[cell_data.list_id]
                    cell_data.field_name = field_info.field_name
                    cell_data.field_source = field_info.source

        # ---------------------------------------------------------------------
        # 5. Excel 워크북 생성
        # ---------------------------------------------------------------------
        print("\n[5] Excel 파일 생성...")
        wb = Workbook()

        # ----- 메인 시트 -----
        if cfg.output.sheets.main:
            ws_main = wb.active
            ws_main.title = sheet_name
            self._create_main_sheet(ws_main, cells_data, row_heights, col_widths, page_result)
            print(f"    [{sheet_name}] 시트 생성 완료")

        # ----- 페이지 정보 시트 -----
        if cfg.output.sheets.page_info and page_result.success:
            ws_page = wb.create_sheet(title=f"{sheet_name}{cfg.output.suffix.page}")
            write_page_info_to_sheet(ws_page, page_result.page_meta)
            print(f"    [{sheet_name}{cfg.output.suffix.page}] 시트 생성 완료")

        # ----- 셀 정보 시트 -----
        if cfg.output.sheets.cell_info:
            ws_cells = wb.create_sheet(title=f"{sheet_name}{cfg.output.suffix.cells}")
            write_cell_styles_to_sheet(ws_cells, cells_data, row_heights, col_widths)
            print(f"    [{sheet_name}{cfg.output.suffix.cells}] 시트 생성 완료")

        # ----- 크기 정보 시트 -----
        if cfg.output.sheets.size_info:
            ws_sizes = wb.create_sheet(title=f"{sheet_name}{cfg.output.suffix.sizes}")
            write_row_col_sizes_to_sheet(ws_sizes, row_heights, col_widths)
            print(f"    [{sheet_name}{cfg.output.suffix.sizes}] 시트 생성 완료")

        # ---------------------------------------------------------------------
        # 6. 저장
        # ---------------------------------------------------------------------
        try:
            wb.save(output_path)
            print("\n" + "=" * 60)
            print(f"저장 완료: {output_path}")
            print("=" * 60)
            return True
        except PermissionError:
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            base, ext = os.path.splitext(output_path)
            alt_path = f"{base}_{timestamp}{ext}"
            wb.save(alt_path)
            print("\n" + "=" * 60)
            print(f"[주의] {output_path}이 열려 있어 다른 이름으로 저장:")
            print(f"저장 완료: {alt_path}")
            print("=" * 60)
            return True

    def _create_main_sheet(self, ws, cells_data, row_heights, col_widths, page_result):
        """메인 표 시트 생성"""
        cfg = self.config

        # 페이지 설정 적용
        if cfg.page.enabled and page_result.success:
            meta = page_result.page_meta

            if cfg.page.margins.enabled:
                margins = PageMargins(
                    left=meta.margin.left / 7200 if cfg.page.margins.left else 0.7,
                    right=meta.margin.right / 7200 if cfg.page.margins.right else 0.7,
                    top=meta.margin.top / 7200 if cfg.page.margins.top else 0.75,
                    bottom=meta.margin.bottom / 7200 if cfg.page.margins.bottom else 0.75,
                )
                ws.page_margins = margins

            if cfg.page.orientation:
                ws.page_setup.orientation = meta.page_size.orientation

            if cfg.page.fit_to_page.enabled:
                ws.page_setup.fitToPage = True
                ws.page_setup.fitToWidth = cfg.page.fit_to_page.width
                ws.page_setup.fitToHeight = cfg.page.fit_to_page.height

        # 테두리 스타일
        border_style = cfg.style.border.style if cfg.style.border.enabled else None
        thin_border = Border(
            left=Side(style=border_style),
            right=Side(style=border_style),
            top=Side(style=border_style),
            bottom=Side(style=border_style)
        ) if border_style else None

        # 셀 보호 스타일
        locked_protection = Protection(locked=True)
        unlocked_protection = Protection(locked=False)

        # 병합 트래킹
        merged_cells = set()

        # 셀 데이터 기록
        for cell_data in cells_data:
            excel_row = cell_data.row + 1
            excel_col = cell_data.col + 1

            if (excel_row, excel_col) in merged_cells:
                continue

            cell = ws.cell(row=excel_row, column=excel_col)
            cell.value = cell_data.text

            if thin_border:
                cell.border = thin_border

            # 스타일 적용
            if cfg.style.enabled:
                self._apply_cell_style(cell, cell_data)

            # 보호 설정
            if cfg.protection.enabled:
                if cell_data.bg_color_rgb:
                    if cfg.protection.lock_rules.with_background:
                        cell.protection = locked_protection
                else:
                    if cfg.protection.lock_rules.without_background:
                        cell.protection = locked_protection
                    else:
                        cell.protection = unlocked_protection

            # 병합 처리
            if cell_data.rowspan > 1 or cell_data.colspan > 1:
                end_row = cell_data.end_row + 1
                end_col = cell_data.end_col + 1

                ws.merge_cells(
                    start_row=excel_row,
                    start_column=excel_col,
                    end_row=end_row,
                    end_column=end_col
                )

                for r in range(excel_row, end_row + 1):
                    for c in range(excel_col, end_col + 1):
                        merged_cells.add((r, c))
                        try:
                            merged_cell = ws.cell(row=r, column=c)
                            if thin_border:
                                merged_cell.border = thin_border

                            if cfg.style.enabled and cfg.style.background.enabled:
                                if cell_data.bg_color_rgb:
                                    rgb = cell_data.bg_color_rgb
                                    hex_color = f"{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}"
                                    merged_cell.fill = PatternFill(
                                        start_color=hex_color,
                                        end_color=hex_color,
                                        fill_type="solid"
                                    )

                            if cfg.protection.enabled:
                                if cell_data.bg_color_rgb and cfg.protection.lock_rules.with_background:
                                    merged_cell.protection = locked_protection
                                elif not cell_data.bg_color_rgb and not cfg.protection.lock_rules.without_background:
                                    merged_cell.protection = unlocked_protection
                        except:
                            pass

        # 크기 설정
        if cfg.size.enabled:
            if cfg.size.col_width.enabled:
                for col_idx, width in enumerate(col_widths):
                    col_letter = get_column_letter(col_idx + 1)
                    char_width = max(cfg.size.col_width.min,
                                    min(width / 700, cfg.size.col_width.max))
                    ws.column_dimensions[col_letter].width = char_width

            if cfg.size.row_height.enabled:
                for row_idx, height in enumerate(row_heights):
                    pt_height = max(cfg.size.row_height.min,
                                   min(height / 100, cfg.size.row_height.max))
                    ws.row_dimensions[row_idx + 1].height = pt_height

        # 시트 보호
        if cfg.protection.enabled:
            ws.protection = SheetProtection(
                sheet=True,
                formatCells=not cfg.protection.allow.format_cells,
                formatColumns=not cfg.protection.allow.format_columns,
                formatRows=not cfg.protection.allow.format_rows,
                insertColumns=not cfg.protection.allow.insert_columns,
                insertRows=not cfg.protection.allow.insert_rows,
                deleteColumns=not cfg.protection.allow.delete_columns,
                deleteRows=not cfg.protection.allow.delete_rows,
            )

    def _apply_cell_style(self, cell, cell_data: CellStyleData):
        """셀 스타일 적용"""
        cfg = self.config.style

        # 배경색
        if cfg.background.enabled and cell_data.bg_color_rgb:
            r, g, b = cell_data.bg_color_rgb
            hex_color = f"{r:02X}{g:02X}{b:02X}"
            cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

        # 글꼴
        if cfg.font.enabled:
            font_kwargs = {}

            if cfg.font.name and cell_data.font_name:
                font_kwargs['name'] = cell_data.font_name

            if cfg.font.size and cell_data.font_size_pt > 0:
                font_kwargs['size'] = cell_data.font_size_pt

            if cfg.font.bold and cell_data.font_bold:
                font_kwargs['bold'] = True

            if cfg.font.italic and cell_data.font_italic:
                font_kwargs['italic'] = True

            if cfg.font.underline and cell_data.font_underline:
                font_kwargs['underline'] = 'single'

            if cfg.font.color and cell_data.font_color_rgb:
                r, g, b = cell_data.font_color_rgb
                font_kwargs['color'] = f"{r:02X}{g:02X}{b:02X}"

            if font_kwargs:
                cell.font = Font(**font_kwargs)

        # 정렬
        if cfg.alignment.enabled:
            h_align_map = {'left': 'left', 'center': 'center', 'right': 'right',
                          'justify': 'justify', 'distribute': 'distributed'}
            v_align_map = {'top': 'top', 'center': 'center', 'bottom': 'bottom'}

            cell.alignment = Alignment(
                horizontal=h_align_map.get(cell_data.align_horizontal, 'left') if cfg.alignment.horizontal else 'left',
                vertical=v_align_map.get(cell_data.align_vertical, 'center') if cfg.alignment.vertical else 'center',
                wrap_text=cfg.alignment.wrap_text
            )


# =============================================================================
# 메인 실행
# =============================================================================

def main():
    parser = argparse.ArgumentParser(description='HWP 표 → Excel 변환')
    parser.add_argument('--config', '-c', type=str, default=None,
                       help='설정 파일 경로 (YAML)')
    parser.add_argument('--output', '-o', type=str, default='C:\\win32hwp\\export_test.xlsx',
                       help='출력 파일 경로')
    parser.add_argument('--sheet', '-s', type=str, default='표',
                       help='메인 시트 이름')

    args = parser.parse_args()

    # 설정 로드
    if args.config:
        print(f"설정 파일 로드: {args.config}")
        config = load_config(args.config)
    else:
        config = get_default_config()

    # 변환 실행
    exporter = HwpToExcelExporter(config)
    exporter.export(args.output, args.sheet)


if __name__ == "__main__":
    main()
