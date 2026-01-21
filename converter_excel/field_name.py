# -*- coding: utf-8 -*-
"""필드 이름 생성 모듈

배경색이 없는 셀에 대해 필드 이름을 생성합니다.

필드 이름 결정 우선순위:
1. 셀에 책갈피가 있으면 → 책갈피 이름
2. A_B 패턴:
   - A: 좌측으로 가서 배경색 있는 셀 (조건: A의 높이 == 현재 셀의 높이)
   - B: 위쪽으로 가서 배경색 있는 셀 (조건: B의 너비 == 현재 셀의 너비)
3. 랜덤 12자리 영문
"""

import random
import string
from dataclasses import dataclass
from typing import Optional, Dict, List, Tuple, Any

try:
    from openpyxl import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


@dataclass
class FieldInfo:
    """필드 정보"""
    list_id: int
    row: int
    col: int
    field_name: str
    source: str  # 'bookmark', 'A_B', 'A', 'B', 'random'
    text: str = ""
    # A/B 셀 정보
    a_text: str = ""
    a_row: int = -1
    a_col: int = -1
    b_text: str = ""
    b_row: int = -1
    b_col: int = -1


def generate_random_field_name(length: int = 12) -> str:
    """랜덤 영문 필드 이름 생성"""
    return ''.join(random.choices(string.ascii_lowercase, k=length))


def get_cell_bookmark(hwp, list_id: int) -> Optional[str]:
    """셀 내부의 책갈피 이름 가져오기

    Args:
        hwp: HWP 객체
        list_id: 셀의 list_id

    Returns:
        책갈피 이름 또는 None
    """
    try:
        # 셀로 이동
        hwp.SetPos(list_id, 0, 0)

        # 셀 내 모든 컨트롤 순회
        ctrl = hwp.HeadCtrl
        while ctrl:
            try:
                ctrl_id = ctrl.CtrlID

                # 책갈피 컨트롤 확인 (bokm: 일반 책갈피, %bmk: 블록 책갈피)
                if ctrl_id in ('bokm', '%bmk'):
                    # 컨트롤의 앵커 위치 확인
                    anchor = ctrl.GetAnchorPos(0)
                    if anchor:
                        ctrl_list_id = anchor.Item("List")
                        if ctrl_list_id == list_id:
                            # 책갈피 이름 가져오기
                            props = ctrl.Properties
                            if props:
                                name = props.Item("Name")
                                if name:
                                    return str(name)
            except:
                pass
            ctrl = ctrl.Next

        return None
    except:
        return None


def find_left_header_cell(
    cells_data: List[Any],
    cells_dict: Dict[int, Any],
    current_cell: Any,
    tolerance: int = 50
) -> Optional[Any]:
    """좌측으로 이동하면서 배경색 있는 셀 찾기 (높이 조건 체크)

    Args:
        cells_data: 모든 셀 데이터 리스트 (CellStyleData)
        cells_dict: list_id -> CellStyleData 매핑
        current_cell: 현재 셀 (CellStyleData)
        tolerance: 높이 비교 허용 오차 (HWPUNIT)

    Returns:
        조건에 맞는 셀 또는 None
    """
    current_row = current_cell.row
    current_end_row = getattr(current_cell, 'end_row', current_row)
    current_height = current_cell.height

    # 좌측 셀들 중 현재 행 범위와 겹치는 셀 찾기
    # (셀의 row ~ end_row 범위가 current_row ~ current_end_row와 겹치면 같은 행으로 간주)
    left_cells = []
    for c in cells_data:
        if c.col >= current_cell.col:
            continue
        cell_end_row = getattr(c, 'end_row', c.row)
        # 범위가 겹치는지 확인
        if c.row <= current_end_row and cell_end_row >= current_row:
            left_cells.append(c)

    # col 내림차순 정렬 (가장 가까운 것부터)
    left_cells.sort(key=lambda c: c.col, reverse=True)

    for cell in left_cells:
        if cell.bg_color_rgb:
            # 높이 조건 체크
            if abs(cell.height - current_height) <= tolerance:
                return cell
            # 높이가 다르면 계속 왼쪽으로
            continue

    return None


def find_top_header_cell(
    cells_data: List[Any],
    cells_dict: Dict[int, Any],
    current_cell: Any,
    tolerance: int = 50
) -> Optional[Any]:
    """위쪽으로 이동하면서 배경색 있는 셀 찾기 (너비 조건 체크)

    Args:
        cells_data: 모든 셀 데이터 리스트 (CellStyleData)
        cells_dict: list_id -> CellStyleData 매핑
        current_cell: 현재 셀 (CellStyleData)
        tolerance: 너비 비교 허용 오차 (HWPUNIT)

    Returns:
        조건에 맞는 셀 또는 None
    """
    current_col = current_cell.col
    current_end_col = getattr(current_cell, 'end_col', current_col)
    current_width = current_cell.width

    # 위쪽 셀들 중 현재 열 범위와 겹치는 셀 찾기
    # (셀의 col ~ end_col 범위가 current_col ~ current_end_col과 겹치면 같은 열로 간주)
    top_cells = []
    for c in cells_data:
        if c.row >= current_cell.row:
            continue
        cell_end_col = getattr(c, 'end_col', c.col)
        # 범위가 겹치는지 확인
        if c.col <= current_end_col and cell_end_col >= current_col:
            top_cells.append(c)

    # row 내림차순 정렬 (가장 가까운 것부터)
    top_cells.sort(key=lambda c: c.row, reverse=True)

    for cell in top_cells:
        if cell.bg_color_rgb:
            # 너비 조건 체크
            if abs(cell.width - current_width) <= tolerance:
                return cell
            # 너비가 다르면 계속 위쪽으로
            continue

    return None


def clean_text_for_field_name(text: str) -> str:
    """텍스트를 필드 이름으로 사용할 수 있게 정리

    - 공백, 특수문자 제거
    - 줄바꿈 제거
    """
    if not text:
        return ""

    # 줄바꿈 제거
    text = text.replace('\r\n', ' ').replace('\r', ' ').replace('\n', ' ')
    # 앞뒤 공백 제거
    text = text.strip()
    # 연속 공백을 하나로
    while '  ' in text:
        text = text.replace('  ', ' ')

    return text


def generate_field_names(
    hwp,
    cells_data: List[Any],
    tolerance: int = 50
) -> List[FieldInfo]:
    """배경색이 없는 모든 셀에 대해 필드 이름 생성

    Args:
        hwp: HWP 객체
        cells_data: 모든 셀 데이터 리스트 (CellStyleData)
        tolerance: 크기 비교 허용 오차 (HWPUNIT)

    Returns:
        FieldInfo 리스트
    """
    # list_id -> CellStyleData 매핑
    cells_dict = {c.list_id: c for c in cells_data}

    # 배경색 없는 셀 필터링
    empty_cells = [c for c in cells_data if not c.bg_color_rgb]

    result = []

    for cell in empty_cells:
        field_info = FieldInfo(
            list_id=cell.list_id,
            row=cell.row,
            col=cell.col,
            field_name="",
            source="",
            text=cell.text
        )

        # 1. 책갈피 확인
        bookmark_name = get_cell_bookmark(hwp, cell.list_id)
        if bookmark_name:
            field_info.field_name = bookmark_name
            field_info.source = "bookmark"
            result.append(field_info)
            continue

        # 2. A_B 패턴
        a_cell = find_left_header_cell(cells_data, cells_dict, cell, tolerance)
        b_cell = find_top_header_cell(cells_data, cells_dict, cell, tolerance)

        a_text = ""
        b_text = ""

        if a_cell:
            a_text = clean_text_for_field_name(a_cell.text)
            field_info.a_text = a_text
            field_info.a_row = a_cell.row
            field_info.a_col = a_cell.col

        if b_cell:
            b_text = clean_text_for_field_name(b_cell.text)
            field_info.b_text = b_text
            field_info.b_row = b_cell.row
            field_info.b_col = b_cell.col

        # 필드 이름 생성
        if a_text and b_text:
            field_name = f"{a_text}_{b_text}"
            field_info.source = "A_B"
        elif a_text:
            field_name = f"{a_text}_"
            field_info.source = "A"
        elif b_text:
            field_name = f"_{b_text}"
            field_info.source = "B"
        else:
            # 3. 랜덤 이름
            field_name = generate_random_field_name()
            field_info.source = "random"

        field_info.field_name = field_name
        result.append(field_info)

    return result


def write_field_info_to_sheet(ws: 'Worksheet', fields: List[FieldInfo]):
    """필드 정보를 엑셀 시트에 기록

    Args:
        ws: openpyxl Worksheet 객체
        fields: FieldInfo 리스트
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl이 필요합니다: pip install openpyxl")

    # 스타일
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 헤더
    headers = [
        "list_id", "row", "col", "field_name", "source",
        "text", "A_text", "A_row", "A_col", "B_text", "B_row", "B_col"
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # 데이터
    sorted_fields = sorted(fields, key=lambda f: (f.row, f.col))

    for row_idx, field in enumerate(sorted_fields, 2):
        values = [
            field.list_id,
            field.row,
            field.col,
            field.field_name,
            field.source,
            field.text[:50] if field.text else "",
            field.a_text,
            field.a_row if field.a_row >= 0 else "",
            field.a_col if field.a_col >= 0 else "",
            field.b_text,
            field.b_row if field.b_row >= 0 else "",
            field.b_col if field.b_col >= 0 else "",
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    # 열 너비 조정
    col_widths = [8, 5, 5, 30, 10, 30, 20, 5, 5, 20, 5, 5]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 인쇄 설정
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def set_cell_field_names(hwp, fields: List[FieldInfo]) -> int:
    """한글 문서의 셀에 필드 이름 설정

    Args:
        hwp: HWP 객체
        fields: FieldInfo 리스트

    Returns:
        성공적으로 설정된 필드 수
    """
    success_count = 0

    for field in fields:
        try:
            # 셀로 이동
            hwp.SetPos(field.list_id, 0, 0)

            # 셀 블록 선택
            hwp.HAction.Run("TableCellBlock")

            # 필드 이름 설정 (option=1: 셀 필드)
            result = hwp.SetCurFieldName(field.field_name, 1, '', '')

            # 선택 해제
            hwp.HAction.Run("Cancel")

            if result:
                success_count += 1
        except Exception as e:
            try:
                hwp.HAction.Run("Cancel")
            except:
                pass

    return success_count


# ============================================================
# 테스트
# ============================================================

if __name__ == "__main__":
    import sys
    import os
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

    from cursor import get_hwp_instance
    from table.table_info import TableInfo
    from table.cell_position import CellPositionCalculator
    from converter_excel.match_cell import extract_cell_style, get_cell_text

    hwp = get_hwp_instance()
    if not hwp:
        print("[오류] 한글이 실행 중이지 않습니다.")
        exit(1)

    table_info = TableInfo(hwp, debug=False)
    if not table_info.is_in_table():
        print("[오류] 커서가 표 안에 있지 않습니다.")
        exit(1)

    print("=== 필드 이름 생성 테스트 ===\n")

    # 셀 위치 계산
    calc = CellPositionCalculator(hwp, debug=False)
    pos_result = calc.calculate()

    # 셀 스타일 추출
    cells_data = []
    for list_id, cell_range in pos_result.cells.items():
        style = extract_cell_style(hwp, list_id)
        style.list_id = list_id
        style.row = cell_range.start_row
        style.col = cell_range.start_col
        style.width = cell_range.end_x - cell_range.start_x
        style.height = cell_range.end_y - cell_range.start_y
        style.text = get_cell_text(hwp, list_id)
        cells_data.append(style)

    # 필드 이름 생성
    fields = generate_field_names(hwp, cells_data)

    print(f"필드 수: {len(fields)}개\n")
    for field in sorted(fields, key=lambda f: (f.row, f.col)):
        print(f"({field.row},{field.col}) {field.field_name} [{field.source}]")
        if field.a_text or field.b_text:
            print(f"    A: {field.a_text}, B: {field.b_text}")

    # 엑셀 저장 테스트
    if HAS_OPENPYXL:
        from openpyxl import Workbook
        import datetime

        wb = Workbook()
        ws = wb.active
        ws.title = "_files"

        write_field_info_to_sheet(ws, fields)

        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath = f"C:\\win32hwp\\test_fields_{timestamp}.xlsx"
        wb.save(filepath)
        print(f"\n저장 완료: {filepath}")
