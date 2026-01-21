# -*- coding: utf-8 -*-
"""페이지 정보를 엑셀에 적용하는 모듈

- write_page_info_to_sheet: 페이지 정보를 시트에 기록
- apply_page_margins_to_excel: 페이지 여백 적용
"""

from typing import TYPE_CHECKING

try:
    from openpyxl.worksheet.page import PageMargins
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

from ..page_meta import HwpPageMeta


# =============================================================================
# 상수 정의
# =============================================================================

HWPUNIT_PER_PT = 100
HWPUNIT_PER_INCH = 7200
HWPUNIT_PER_CM = 7200 / 2.54  # 약 2834.6


# =============================================================================
# 스타일 헬퍼
# =============================================================================

def _get_header_style():
    """헤더 셀 스타일 반환"""
    if not HAS_OPENPYXL:
        return None, None, None

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    return header_font, header_fill, thin_border


# =============================================================================
# 페이지 관련 함수
# =============================================================================

def write_page_info_to_sheet(ws: 'Worksheet', page_meta: HwpPageMeta,
                              sheet_title: str = "페이지 설정"):
    """페이지 정보를 엑셀 시트에 기록

    Args:
        ws: openpyxl Worksheet 객체
        page_meta: HwpPageMeta 객체
        sheet_title: 시트 제목
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl이 필요합니다: pip install openpyxl")

    header_font, header_fill, thin_border = _get_header_style()

    # 헤더
    headers = ["항목", "값 (HWPUNIT)", "값 (cm)", "값 (inch)", "값 (pt)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # 데이터
    m = page_meta.margin
    ps = page_meta.page_size

    data = [
        # 용지 크기
        ("용지 너비", ps.width, ps.width / HWPUNIT_PER_CM, ps.width / HWPUNIT_PER_INCH, ps.width / 100),
        ("용지 높이", ps.height, ps.height / HWPUNIT_PER_CM, ps.height / HWPUNIT_PER_INCH, ps.height / 100),
        ("용지 방향", ps.orientation, "", "", ""),
        ("", "", "", "", ""),
        # 여백
        ("왼쪽 여백", m.left, m.left / HWPUNIT_PER_CM, m.left / HWPUNIT_PER_INCH, m.left / 100),
        ("오른쪽 여백", m.right, m.right / HWPUNIT_PER_CM, m.right / HWPUNIT_PER_INCH, m.right / 100),
        ("위쪽 여백", m.top, m.top / HWPUNIT_PER_CM, m.top / HWPUNIT_PER_INCH, m.top / 100),
        ("아래쪽 여백", m.bottom, m.bottom / HWPUNIT_PER_CM, m.bottom / HWPUNIT_PER_INCH, m.bottom / 100),
        ("머리말", m.header, m.header / HWPUNIT_PER_CM, m.header / HWPUNIT_PER_INCH, m.header / 100),
        ("꼬리말", m.footer, m.footer / HWPUNIT_PER_CM, m.footer / HWPUNIT_PER_INCH, m.footer / 100),
        ("제본 여백", m.gutter, m.gutter / HWPUNIT_PER_CM, m.gutter / HWPUNIT_PER_INCH, m.gutter / 100),
        ("", "", "", "", ""),
        # 본문 영역
        ("본문 너비", page_meta.content_width,
         page_meta.content_width / HWPUNIT_PER_CM,
         page_meta.content_width / HWPUNIT_PER_INCH,
         page_meta.content_width / 100),
        ("본문 높이", page_meta.content_height,
         page_meta.content_height / HWPUNIT_PER_CM,
         page_meta.content_height / HWPUNIT_PER_INCH,
         page_meta.content_height / 100),
    ]

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)

            # 숫자 포맷
            if isinstance(value, float):
                cell.value = round(value, 2)
                cell.number_format = '0.00'
            else:
                cell.value = value

            cell.border = thin_border
            if col_idx == 1:
                cell.alignment = Alignment(horizontal='left')
            else:
                cell.alignment = Alignment(horizontal='right')

    # 열 너비 조정
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12

    # 인쇄 설정
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(
        left=0.5, right=0.5, top=0.5, bottom=0.5,
        header=0.3, footer=0.3
    )


def apply_page_margins_to_excel(ws: 'Worksheet', page_meta: HwpPageMeta):
    """한글 페이지 여백을 엑셀 워크시트에 적용

    Args:
        ws: openpyxl Worksheet 객체
        page_meta: HwpPageMeta 객체
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl이 필요합니다: pip install openpyxl")

    ws.page_margins = PageMargins(
        left=page_meta.margin.left / HWPUNIT_PER_INCH,
        right=page_meta.margin.right / HWPUNIT_PER_INCH,
        top=page_meta.margin.top / HWPUNIT_PER_INCH,
        bottom=page_meta.margin.bottom / HWPUNIT_PER_INCH,
        header=page_meta.margin.header / HWPUNIT_PER_INCH,
        footer=page_meta.margin.footer / HWPUNIT_PER_INCH,
    )

    # 페이지 방향
    ws.page_setup.orientation = page_meta.page_size.orientation
