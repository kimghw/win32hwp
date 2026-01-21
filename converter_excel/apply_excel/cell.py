# -*- coding: utf-8 -*-
"""셀 스타일 정보를 엑셀에 적용하는 모듈

- write_cell_styles_to_sheet: 셀 스타일 정보를 시트에 기록
- write_row_col_sizes_to_sheet: 행/열 크기 정보를 시트에 기록
- apply_cell_style_to_excel_cell: 셀 스타일을 엑셀 셀에 적용
"""

from typing import List, TYPE_CHECKING

try:
    from openpyxl.worksheet.page import PageMargins
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

from ..extract_data_hwp.cell import CellStyleData


# =============================================================================
# 상수 정의
# =============================================================================

HWPUNIT_PER_CM = 7200 / 2.54  # 약 2834.6


# =============================================================================
# 스타일 헬퍼
# =============================================================================

def _get_header_style():
    """헤더 셀 스타일 반환"""
    if not HAS_OPENPYXL:
        return None, None, None

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    return header_font, header_fill, thin_border


def _get_thin_border():
    """얇은 테두리 스타일 반환"""
    if not HAS_OPENPYXL:
        return None

    return Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )


# =============================================================================
# 셀 스타일 관련 함수
# =============================================================================

def write_cell_styles_to_sheet(ws: 'Worksheet', cells: List[CellStyleData],
                                row_heights: List[int] = None,
                                col_widths: List[int] = None):
    """셀 스타일 정보를 엑셀 시트에 기록

    Args:
        ws: openpyxl Worksheet 객체
        cells: CellStyleData 리스트
        row_heights: 행 높이 리스트 (HWPUNIT)
        col_widths: 열 너비 리스트 (HWPUNIT)
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl이 필요합니다: pip install openpyxl")

    header_font, header_fill, thin_border = _get_header_style()

    # 헤더
    headers = [
        # 위치 정보
        "list_id", "row", "col", "end_row", "end_col", "rowspan", "colspan",
        # 물리 좌표
        "x", "y", "width", "height", "width_pt", "height_pt",
        # 배경색
        "bg_color",
        # 셀 내부 여백 (pt)
        "margin_L", "margin_R", "margin_T", "margin_B",
        # 글꼴 기본
        "font_name", "font_size", "bold", "italic", "font_color",
        # 글꼴 장식
        "underline", "strikeout", "outline", "shadow", "emboss", "engrave",
        "superscript", "subscript",
        # 자간/장평/줄간격
        "char_spacing", "char_ratio", "line_spacing",
        # 정렬
        "align_h", "align_v",
        # 테두리
        "border_L", "border_R", "border_T", "border_B",
        # 텍스트
        "text",
        # 필드
        "field_name", "field_source"
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # 셀 데이터
    sorted_cells = sorted(cells, key=lambda c: (c.row, c.col))

    for row_idx, cell_data in enumerate(sorted_cells, 2):
        values = [
            # 위치 정보
            cell_data.list_id,
            cell_data.row,
            cell_data.col,
            cell_data.end_row,
            cell_data.end_col,
            cell_data.rowspan,
            cell_data.colspan,
            # 물리 좌표
            cell_data.x,
            cell_data.y,
            cell_data.width,
            cell_data.height,
            round(cell_data.width / 100, 1),   # width_pt
            round(cell_data.height / 100, 1),  # height_pt
            # 배경색
            f"#{cell_data.bg_color_rgb[0]:02X}{cell_data.bg_color_rgb[1]:02X}{cell_data.bg_color_rgb[2]:02X}" if cell_data.bg_color_rgb else "",
            # 셀 내부 여백 (pt)
            round(cell_data.margin_left / 100, 1),
            round(cell_data.margin_right / 100, 1),
            round(cell_data.margin_top / 100, 1),
            round(cell_data.margin_bottom / 100, 1),
            # 글꼴 기본
            cell_data.font_name or "",
            cell_data.font_size_pt,
            "Y" if cell_data.font_bold else "",
            "Y" if cell_data.font_italic else "",
            f"#{cell_data.font_color_rgb[0]:02X}{cell_data.font_color_rgb[1]:02X}{cell_data.font_color_rgb[2]:02X}" if cell_data.font_color_rgb else "",
            # 글꼴 장식
            cell_data.font_underline if cell_data.font_underline else "",
            cell_data.font_strikeout if cell_data.font_strikeout else "",
            cell_data.font_outline if cell_data.font_outline else "",
            cell_data.font_shadow if cell_data.font_shadow else "",
            "Y" if cell_data.font_emboss else "",
            "Y" if cell_data.font_engrave else "",
            "Y" if cell_data.font_superscript else "",
            "Y" if cell_data.font_subscript else "",
            # 자간/장평/줄간격
            cell_data.char_spacing,
            cell_data.char_ratio,
            cell_data.line_spacing,
            # 정렬
            cell_data.align_horizontal,
            cell_data.align_vertical,
            # 테두리 (type/width)
            f"{cell_data.border_left_type}/{cell_data.border_left_width}",
            f"{cell_data.border_right_type}/{cell_data.border_right_width}",
            f"{cell_data.border_top_type}/{cell_data.border_top_width}",
            f"{cell_data.border_bottom_type}/{cell_data.border_bottom_width}",
            # 텍스트
            cell_data.text[:50] if cell_data.text else "",  # 50자 제한
            # 필드
            cell_data.field_name or "",
            cell_data.field_source or ""
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

            # 배경색이 있는 셀 표시
            if col_idx == 14 and cell_data.bg_color_rgb:  # bg_color 열
                r, g, b = cell_data.bg_color_rgb
                cell.fill = PatternFill(start_color=f"{r:02X}{g:02X}{b:02X}",
                                        end_color=f"{r:02X}{g:02X}{b:02X}",
                                        fill_type="solid")

    # 열 너비 조정
    col_widths_excel = [
        8, 5, 5, 7, 7, 7, 7,     # 위치 (7개)
        8, 8, 8, 8, 8, 8,        # 물리 좌표 (6개)
        10,                       # 배경색 (1개)
        7, 7, 7, 7,              # 여백 (4개)
        12, 8, 5, 5, 10,         # 글꼴 기본: name, size, bold, italic, color (5개)
        8, 8, 7, 7, 7, 7, 8, 8,  # 글꼴 장식: underline, strikeout, outline, shadow, emboss, engrave, superscript, subscript (8개)
        10, 10, 10,              # 자간/장평/줄간격 (3개)
        8, 8,                     # 정렬 (2개)
        8, 8, 8, 8,              # 테두리 (4개)
        25,                       # 텍스트 (1개)
        25, 10                    # 필드: field_name, field_source (2개)
    ]
    for col_idx, width in enumerate(col_widths_excel, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 인쇄 설정
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(
        left=0.3, right=0.3, top=0.3, bottom=0.3,
        header=0.2, footer=0.2
    )


def write_row_col_sizes_to_sheet(ws: 'Worksheet', row_heights: List[int], col_widths: List[int]):
    """행 높이와 열 너비를 별도 시트에 기록

    Args:
        ws: openpyxl Worksheet 객체
        row_heights: 행 높이 리스트 (HWPUNIT)
        col_widths: 열 너비 리스트 (HWPUNIT)
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl이 필요합니다: pip install openpyxl")

    header_font, header_fill, thin_border = _get_header_style()

    # 행 높이 섹션
    ws.cell(row=1, column=1, value="행 높이").font = header_font
    headers_row = ["행 번호", "HWPUNIT", "pt", "cm"]
    for col, header in enumerate(headers_row, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for idx, height in enumerate(row_heights):
        row = idx + 3
        ws.cell(row=row, column=1, value=idx + 1).border = thin_border
        ws.cell(row=row, column=2, value=height).border = thin_border
        ws.cell(row=row, column=3, value=round(height / 100, 2)).border = thin_border
        ws.cell(row=row, column=4, value=round(height / HWPUNIT_PER_CM, 2)).border = thin_border

    # 열 너비 섹션 (오른쪽에)
    start_col = 6
    ws.cell(row=1, column=start_col, value="열 너비").font = header_font
    headers_col = ["열 번호", "HWPUNIT", "pt", "cm", "문자(chars)"]
    for col, header in enumerate(headers_col, start_col):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for idx, width in enumerate(col_widths):
        row = idx + 3
        ws.cell(row=row, column=start_col, value=idx + 1).border = thin_border
        ws.cell(row=row, column=start_col + 1, value=width).border = thin_border
        ws.cell(row=row, column=start_col + 2, value=round(width / 100, 2)).border = thin_border
        ws.cell(row=row, column=start_col + 3, value=round(width / HWPUNIT_PER_CM, 2)).border = thin_border
        ws.cell(row=row, column=start_col + 4, value=round(width / 700, 2)).border = thin_border

    # 열 너비 조정
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 8
    ws.column_dimensions['I'].width = 8
    ws.column_dimensions['J'].width = 10

    # 인쇄 설정
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def apply_cell_style_to_excel_cell(ws: 'Worksheet', excel_row: int, excel_col: int,
                                    cell_data: CellStyleData):
    """셀 스타일을 엑셀 셀에 적용

    Args:
        ws: Worksheet 객체
        excel_row: 엑셀 행 번호 (1-based)
        excel_col: 엑셀 열 번호 (1-based)
        cell_data: CellStyleData 객체
    """
    if not HAS_OPENPYXL:
        return

    cell = ws.cell(row=excel_row, column=excel_col)

    # 배경색
    if cell_data.bg_color_rgb:
        r, g, b = cell_data.bg_color_rgb
        hex_color = f"{r:02X}{g:02X}{b:02X}"
        cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

    # 글꼴
    font_kwargs = {}
    if cell_data.font_name:
        font_kwargs['name'] = cell_data.font_name
    if cell_data.font_size_pt > 0:
        font_kwargs['size'] = cell_data.font_size_pt
    if cell_data.font_bold:
        font_kwargs['bold'] = True
    if cell_data.font_italic:
        font_kwargs['italic'] = True
    if cell_data.font_underline:
        font_kwargs['underline'] = 'single'
    if cell_data.font_color_rgb:
        r, g, b = cell_data.font_color_rgb
        font_kwargs['color'] = f"{r:02X}{g:02X}{b:02X}"

    if font_kwargs:
        cell.font = Font(**font_kwargs)

    # 정렬
    h_align_map = {'left': 'left', 'center': 'center', 'right': 'right',
                   'justify': 'justify', 'distribute': 'distributed'}
    v_align_map = {'top': 'top', 'center': 'center', 'bottom': 'bottom'}

    cell.alignment = Alignment(
        horizontal=h_align_map.get(cell_data.align_horizontal, 'left'),
        vertical=v_align_map.get(cell_data.align_vertical, 'center'),
        wrap_text=True
    )
