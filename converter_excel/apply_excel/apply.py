# -*- coding: utf-8 -*-
"""엑셀 데이터 적용 및 저장 통합 모듈

한글에서 추출한 페이지, 셀, 필드 정보를 엑셀 파일에 적용하고 저장합니다.

주요 기능:
- create_main_sheet: 메인 표 시트 생성
- apply_to_excel: 추출 데이터를 엑셀로 저장 (통합 함수)
"""

import os
import datetime
from typing import List, Dict, Any, TYPE_CHECKING

try:
    from openpyxl import Workbook
    from openpyxl.worksheet.page import PageMargins
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Protection
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.protection import SheetProtection
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

# 데이터 클래스 임포트 (extract_data_hwp에서)
from ..extract_data_hwp.cell import CellStyleData
from ..extract_data_hwp.page import PageMatchResult
from ..extract_data_hwp.field import FieldInfo
from ..page_meta import HwpPageMeta
from ..config import ExportConfig, get_default_config

# 하위 모듈 임포트
from .page import write_page_info_to_sheet
from .cell import write_cell_styles_to_sheet, write_row_col_sizes_to_sheet
from .field import write_field_info_to_sheet


# =============================================================================
# 상수 정의
# =============================================================================

HWPUNIT_PER_PT = 100
HWPUNIT_PER_INCH = 7200
HWPUNIT_PER_CM = 7200 / 2.54  # 약 2834.6


# =============================================================================
# 메인 시트 생성 함수
# =============================================================================

def create_main_sheet(ws: 'Worksheet',
                      cells_data: List[CellStyleData],
                      row_heights: List[int],
                      col_widths: List[int],
                      page_result: PageMatchResult,
                      config: ExportConfig = None):
    """메인 표 시트 생성

    Args:
        ws: Worksheet 객체
        cells_data: 셀 데이터 리스트
        row_heights: 행 높이 리스트 (HWPUNIT)
        col_widths: 열 너비 리스트 (HWPUNIT)
        page_result: PageMatchResult 객체
        config: ExportConfig 설정 객체
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl이 필요합니다: pip install openpyxl")

    cfg = config or get_default_config()

    # 페이지 설정 적용
    if cfg.page.enabled and page_result.success:
        meta = page_result.page_meta

        if cfg.page.margins.enabled:
            margins = PageMargins(
                left=meta.margin.left / HWPUNIT_PER_INCH if cfg.page.margins.left else 0.7,
                right=meta.margin.right / HWPUNIT_PER_INCH if cfg.page.margins.right else 0.7,
                top=meta.margin.top / HWPUNIT_PER_INCH if cfg.page.margins.top else 0.75,
                bottom=meta.margin.bottom / HWPUNIT_PER_INCH if cfg.page.margins.bottom else 0.75,
            )
            ws.page_margins = margins

        if cfg.page.orientation:
            ws.page_setup.orientation = meta.page_size.orientation

        if cfg.page.fit_to_page.enabled:
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToWidth = cfg.page.fit_to_page.width
            ws.page_setup.fitToHeight = cfg.page.fit_to_page.height

    # 테두리 스타일
    border_style = cfg.style.border.style if cfg.style.border.enabled else None
    thin_border = Border(
        left=Side(style=border_style),
        right=Side(style=border_style),
        top=Side(style=border_style),
        bottom=Side(style=border_style)
    ) if border_style else None

    # 셀 보호 스타일
    locked_protection = Protection(locked=True)
    unlocked_protection = Protection(locked=False)

    # 병합 트래킹
    merged_cells = set()

    # 셀 데이터 기록
    for cell_data in cells_data:
        excel_row = cell_data.row + 1
        excel_col = cell_data.col + 1

        if (excel_row, excel_col) in merged_cells:
            continue

        cell = ws.cell(row=excel_row, column=excel_col)
        cell.value = cell_data.text

        if thin_border:
            cell.border = thin_border

        # 스타일 적용
        if cfg.style.enabled:
            _apply_cell_style(cell, cell_data, cfg)

        # 보호 설정
        if cfg.protection.enabled:
            if cell_data.bg_color_rgb:
                if cfg.protection.lock_rules.with_background:
                    cell.protection = locked_protection
            else:
                if cfg.protection.lock_rules.without_background:
                    cell.protection = locked_protection
                else:
                    cell.protection = unlocked_protection

        # 병합 처리
        if cell_data.rowspan > 1 or cell_data.colspan > 1:
            end_row = cell_data.end_row + 1
            end_col = cell_data.end_col + 1

            ws.merge_cells(
                start_row=excel_row,
                start_column=excel_col,
                end_row=end_row,
                end_column=end_col
            )

            for r in range(excel_row, end_row + 1):
                for c in range(excel_col, end_col + 1):
                    merged_cells.add((r, c))
                    try:
                        merged_cell = ws.cell(row=r, column=c)
                        if thin_border:
                            merged_cell.border = thin_border

                        if cfg.style.enabled and cfg.style.background.enabled:
                            if cell_data.bg_color_rgb:
                                rgb = cell_data.bg_color_rgb
                                hex_color = f"{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}"
                                merged_cell.fill = PatternFill(
                                    start_color=hex_color,
                                    end_color=hex_color,
                                    fill_type="solid"
                                )

                        if cfg.protection.enabled:
                            if cell_data.bg_color_rgb and cfg.protection.lock_rules.with_background:
                                merged_cell.protection = locked_protection
                            elif not cell_data.bg_color_rgb and not cfg.protection.lock_rules.without_background:
                                merged_cell.protection = unlocked_protection
                    except:
                        pass

    # 크기 설정
    if cfg.size.enabled:
        if cfg.size.col_width.enabled:
            for col_idx, width in enumerate(col_widths):
                col_letter = get_column_letter(col_idx + 1)
                char_width = max(cfg.size.col_width.min,
                                min(width / 700, cfg.size.col_width.max))
                ws.column_dimensions[col_letter].width = char_width

        if cfg.size.row_height.enabled:
            for row_idx, height in enumerate(row_heights):
                pt_height = max(cfg.size.row_height.min,
                               min(height / 100, cfg.size.row_height.max))
                ws.row_dimensions[row_idx + 1].height = pt_height

    # 시트 보호
    if cfg.protection.enabled:
        ws.protection = SheetProtection(
            sheet=True,
            formatCells=not cfg.protection.allow.format_cells,
            formatColumns=not cfg.protection.allow.format_columns,
            formatRows=not cfg.protection.allow.format_rows,
            insertColumns=not cfg.protection.allow.insert_columns,
            insertRows=not cfg.protection.allow.insert_rows,
            deleteColumns=not cfg.protection.allow.delete_columns,
            deleteRows=not cfg.protection.allow.delete_rows,
        )


def _apply_cell_style(cell, cell_data: CellStyleData, config: ExportConfig):
    """셀 스타일 적용 (내부 함수)

    Args:
        cell: openpyxl Cell 객체
        cell_data: CellStyleData 객체
        config: ExportConfig 설정 객체
    """
    cfg = config.style

    # 배경색
    if cfg.background.enabled and cell_data.bg_color_rgb:
        r, g, b = cell_data.bg_color_rgb
        hex_color = f"{r:02X}{g:02X}{b:02X}"
        cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

    # 글꼴
    if cfg.font.enabled:
        font_kwargs = {}

        if cfg.font.name and cell_data.font_name:
            font_kwargs['name'] = cell_data.font_name

        if cfg.font.size and cell_data.font_size_pt > 0:
            font_kwargs['size'] = cell_data.font_size_pt

        if cfg.font.bold and cell_data.font_bold:
            font_kwargs['bold'] = True

        if cfg.font.italic and cell_data.font_italic:
            font_kwargs['italic'] = True

        if cfg.font.underline and cell_data.font_underline:
            font_kwargs['underline'] = 'single'

        if cfg.font.color and cell_data.font_color_rgb:
            r, g, b = cell_data.font_color_rgb
            font_kwargs['color'] = f"{r:02X}{g:02X}{b:02X}"

        if font_kwargs:
            cell.font = Font(**font_kwargs)

    # 정렬
    if cfg.alignment.enabled:
        h_align_map = {'left': 'left', 'center': 'center', 'right': 'right',
                      'justify': 'justify', 'distribute': 'distributed'}
        v_align_map = {'top': 'top', 'center': 'center', 'bottom': 'bottom'}

        cell.alignment = Alignment(
            horizontal=h_align_map.get(cell_data.align_horizontal, 'left') if cfg.alignment.horizontal else 'left',
            vertical=v_align_map.get(cell_data.align_vertical, 'center') if cfg.alignment.vertical else 'center',
            wrap_text=cfg.alignment.wrap_text
        )


# =============================================================================
# 통합 저장 함수
# =============================================================================

def apply_to_excel(extracted_data: Dict[str, Any],
                   output_path: str,
                   sheet_name: str = "표",
                   config: ExportConfig = None) -> bool:
    """추출된 데이터를 엑셀 파일로 저장

    Args:
        extracted_data: 추출된 데이터 딕셔너리
            - 'cells': List[CellStyleData] - 셀 데이터 리스트
            - 'row_heights': List[int] - 행 높이 리스트 (HWPUNIT)
            - 'col_widths': List[int] - 열 너비 리스트 (HWPUNIT)
            - 'page_result': PageMatchResult - 페이지 정보
            - 'fields': List[FieldInfo] - 필드 정보 (선택)
        output_path: 출력 파일 경로
        sheet_name: 메인 시트 이름
        config: ExportConfig 설정 객체

    Returns:
        성공 여부
    """
    if not HAS_OPENPYXL:
        print("[오류] openpyxl이 필요합니다: pip install openpyxl")
        return False

    cfg = config or get_default_config()

    # 데이터 추출
    cells_data = extracted_data.get('cells', [])
    row_heights = extracted_data.get('row_heights', [])
    col_widths = extracted_data.get('col_widths', [])
    page_result = extracted_data.get('page_result')
    fields = extracted_data.get('fields', [])

    # 워크북 생성
    wb = Workbook()

    # ----- 메인 시트 -----
    if cfg.output.sheets.main:
        ws_main = wb.active
        ws_main.title = sheet_name
        create_main_sheet(ws_main, cells_data, row_heights, col_widths, page_result, cfg)
        print(f"    [{sheet_name}] 시트 생성 완료")

    # ----- 페이지 정보 시트 -----
    if cfg.output.sheets.page_info and page_result and page_result.success:
        ws_page = wb.create_sheet(title=f"{sheet_name}{cfg.output.suffix.page}")
        write_page_info_to_sheet(ws_page, page_result.page_meta)
        print(f"    [{sheet_name}{cfg.output.suffix.page}] 시트 생성 완료")

    # ----- 셀 정보 시트 -----
    if cfg.output.sheets.cell_info:
        ws_cells = wb.create_sheet(title=f"{sheet_name}{cfg.output.suffix.cells}")
        write_cell_styles_to_sheet(ws_cells, cells_data, row_heights, col_widths)
        print(f"    [{sheet_name}{cfg.output.suffix.cells}] 시트 생성 완료")

    # ----- 크기 정보 시트 -----
    if cfg.output.sheets.size_info:
        ws_sizes = wb.create_sheet(title=f"{sheet_name}{cfg.output.suffix.sizes}")
        write_row_col_sizes_to_sheet(ws_sizes, row_heights, col_widths)
        print(f"    [{sheet_name}{cfg.output.suffix.sizes}] 시트 생성 완료")

    # ----- 필드 정보 시트 (추가) -----
    if fields and cfg.field.enabled:
        ws_fields = wb.create_sheet(title=f"{sheet_name}_fields")
        write_field_info_to_sheet(ws_fields, fields)
        print(f"    [{sheet_name}_fields] 시트 생성 완료")

    # 저장
    try:
        wb.save(output_path)
        print(f"\n저장 완료: {output_path}")
        return True
    except PermissionError:
        # 파일이 열려 있는 경우 타임스탬프 추가
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        base, ext = os.path.splitext(output_path)
        alt_path = f"{base}_{timestamp}{ext}"
        wb.save(alt_path)
        print(f"\n[주의] {output_path}이 열려 있어 다른 이름으로 저장:")
        print(f"저장 완료: {alt_path}")
        return True
    except Exception as e:
        print(f"\n[오류] 저장 실패: {e}")
        return False


# =============================================================================
# 테스트
# =============================================================================

if __name__ == "__main__":
    print("=== apply_excel 모듈 테스트 ===")
    print(f"HAS_OPENPYXL: {HAS_OPENPYXL}")

    if HAS_OPENPYXL:
        # 간단한 테스트
        wb = Workbook()
        ws = wb.active
        ws.title = "테스트"
        ws.cell(row=1, column=1, value="테스트 셀")

        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        test_path = f"C:\\win32hwp\\apply_excel_test_{timestamp}.xlsx"
        wb.save(test_path)
        print(f"테스트 파일 저장: {test_path}")
