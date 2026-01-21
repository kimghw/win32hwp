# -*- coding: utf-8 -*-
"""필드 정보를 엑셀에 적용하는 모듈

- write_field_info_to_sheet: 필드 정보를 시트에 기록
"""

from typing import List, TYPE_CHECKING

try:
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

from ..extract_data_hwp.field import FieldInfo


# =============================================================================
# 스타일 헬퍼
# =============================================================================

def _get_header_style():
    """헤더 셀 스타일 반환"""
    if not HAS_OPENPYXL:
        return None, None, None

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    return header_font, header_fill, thin_border


# =============================================================================
# 필드 관련 함수
# =============================================================================

def write_field_info_to_sheet(ws: 'Worksheet', fields: List[FieldInfo]):
    """필드 정보를 엑셀 시트에 기록

    Args:
        ws: openpyxl Worksheet 객체
        fields: FieldInfo 리스트
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl이 필요합니다: pip install openpyxl")

    header_font, header_fill, thin_border = _get_header_style()

    # 헤더
    headers = [
        "list_id", "row", "col", "field_name", "source",
        "text", "A_text", "A_row", "A_col", "B_text", "B_row", "B_col"
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # 데이터
    sorted_fields = sorted(fields, key=lambda f: (f.row, f.col))

    for row_idx, field in enumerate(sorted_fields, 2):
        values = [
            field.list_id,
            field.row,
            field.col,
            field.field_name,
            field.source,
            field.text[:50] if field.text else "",
            field.a_text,
            field.a_row if field.a_row >= 0 else "",
            field.a_col if field.a_col >= 0 else "",
            field.b_text,
            field.b_row if field.b_row >= 0 else "",
            field.b_col if field.b_col >= 0 else "",
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    # 열 너비 조정
    col_widths = [8, 5, 5, 30, 10, 30, 20, 5, 5, 20, 5, 5]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 인쇄 설정
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
