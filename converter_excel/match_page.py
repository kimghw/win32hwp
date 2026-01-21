# -*- coding: utf-8 -*-
"""한글 페이지 설정 → 엑셀 시트 저장 모듈

페이지 여백, 용지 크기, 방향 등의 정보를 엑셀 시트에 저장합니다.
"""

from dataclasses import dataclass, field
from typing import Optional, Dict, List, Any

try:
    from openpyxl import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.worksheet.page import PageMargins
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from .page_meta import Unit, PageMargin, PageSize, HwpPageMeta, get_hwp_page_meta


@dataclass
class PageMatchResult:
    """페이지 매칭 결과"""
    page_meta: HwpPageMeta = None
    success: bool = False
    error: str = None


def extract_page_info(hwp) -> PageMatchResult:
    """한글에서 페이지 정보 추출

    Args:
        hwp: HWP 객체

    Returns:
        PageMatchResult 객체
    """
    result = PageMatchResult()

    try:
        meta = HwpPageMeta()

        act = hwp.CreateAction("PageSetup")
        pset = act.CreateSet()
        act.GetDefault(pset)
        page_def = pset.Item("PageDef")

        # 용지 크기
        meta.page_size.width = page_def.Item("PaperWidth")
        meta.page_size.height = page_def.Item("PaperHeight")
        meta.page_size.orientation = 'landscape' if page_def.Item("Landscape") else 'portrait'

        # 여백
        meta.margin.left = page_def.Item("LeftMargin")
        meta.margin.right = page_def.Item("RightMargin")
        meta.margin.top = page_def.Item("TopMargin")
        meta.margin.bottom = page_def.Item("BottomMargin")
        meta.margin.header = page_def.Item("HeaderLen")
        meta.margin.footer = page_def.Item("FooterLen")
        meta.margin.gutter = page_def.Item("GutterLen")

        # 본문 영역 계산
        meta.calculate_content_area()

        result.page_meta = meta
        result.success = True

    except Exception as e:
        result.error = str(e)
        result.success = False

    return result


def apply_page_margins_to_excel(ws: 'Worksheet', page_meta: HwpPageMeta):
    """한글 페이지 여백을 엑셀 워크시트에 적용

    Args:
        ws: openpyxl Worksheet 객체
        page_meta: HwpPageMeta 객체
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl이 필요합니다: pip install openpyxl")

    # HWPUNIT -> 인치 변환
    HWPUNIT_PER_INCH = 7200

    ws.page_margins = PageMargins(
        left=page_meta.margin.left / HWPUNIT_PER_INCH,
        right=page_meta.margin.right / HWPUNIT_PER_INCH,
        top=page_meta.margin.top / HWPUNIT_PER_INCH,
        bottom=page_meta.margin.bottom / HWPUNIT_PER_INCH,
        header=page_meta.margin.header / HWPUNIT_PER_INCH,
        footer=page_meta.margin.footer / HWPUNIT_PER_INCH,
    )

    # 페이지 방향
    ws.page_setup.orientation = page_meta.page_size.orientation


def write_page_info_to_sheet(ws: 'Worksheet', page_meta: HwpPageMeta,
                              sheet_title: str = "페이지 설정"):
    """페이지 정보를 엑셀 시트에 기록

    Args:
        ws: openpyxl Worksheet 객체
        page_meta: HwpPageMeta 객체
        sheet_title: 시트 제목
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl이 필요합니다: pip install openpyxl")

    # 단위 변환 상수
    HWPUNIT_PER_INCH = 7200
    HWPUNIT_PER_CM = 7200 / 2.54

    # 스타일
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 헤더
    headers = ["항목", "값 (HWPUNIT)", "값 (cm)", "값 (inch)", "값 (pt)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # 데이터
    m = page_meta.margin
    ps = page_meta.page_size

    data = [
        # 용지 크기
        ("용지 너비", ps.width, ps.width / HWPUNIT_PER_CM, ps.width / HWPUNIT_PER_INCH, ps.width / 100),
        ("용지 높이", ps.height, ps.height / HWPUNIT_PER_CM, ps.height / HWPUNIT_PER_INCH, ps.height / 100),
        ("용지 방향", ps.orientation, "", "", ""),
        ("", "", "", "", ""),
        # 여백
        ("왼쪽 여백", m.left, m.left / HWPUNIT_PER_CM, m.left / HWPUNIT_PER_INCH, m.left / 100),
        ("오른쪽 여백", m.right, m.right / HWPUNIT_PER_CM, m.right / HWPUNIT_PER_INCH, m.right / 100),
        ("위쪽 여백", m.top, m.top / HWPUNIT_PER_CM, m.top / HWPUNIT_PER_INCH, m.top / 100),
        ("아래쪽 여백", m.bottom, m.bottom / HWPUNIT_PER_CM, m.bottom / HWPUNIT_PER_INCH, m.bottom / 100),
        ("머리말", m.header, m.header / HWPUNIT_PER_CM, m.header / HWPUNIT_PER_INCH, m.header / 100),
        ("꼬리말", m.footer, m.footer / HWPUNIT_PER_CM, m.footer / HWPUNIT_PER_INCH, m.footer / 100),
        ("제본 여백", m.gutter, m.gutter / HWPUNIT_PER_CM, m.gutter / HWPUNIT_PER_INCH, m.gutter / 100),
        ("", "", "", "", ""),
        # 본문 영역
        ("본문 너비", page_meta.content_width,
         page_meta.content_width / HWPUNIT_PER_CM,
         page_meta.content_width / HWPUNIT_PER_INCH,
         page_meta.content_width / 100),
        ("본문 높이", page_meta.content_height,
         page_meta.content_height / HWPUNIT_PER_CM,
         page_meta.content_height / HWPUNIT_PER_INCH,
         page_meta.content_height / 100),
    ]

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)

            # 숫자 포맷
            if isinstance(value, float):
                cell.value = round(value, 2)
                cell.number_format = '0.00'
            else:
                cell.value = value

            cell.border = thin_border
            if col_idx == 1:
                cell.alignment = Alignment(horizontal='left')
            else:
                cell.alignment = Alignment(horizontal='right')

    # 열 너비 조정
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12

    # 인쇄 설정
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(
        left=0.5, right=0.5, top=0.5, bottom=0.5,
        header=0.3, footer=0.3
    )


def save_page_info_to_excel(hwp, filepath: str, sheet_name: str = "_page") -> str:
    """한글 페이지 정보를 엑셀 파일로 저장 (단독 실행용)

    Args:
        hwp: HWP 객체
        filepath: 저장할 엑셀 파일 경로
        sheet_name: 시트 이름

    Returns:
        저장된 파일 경로
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl이 필요합니다: pip install openpyxl")

    result = extract_page_info(hwp)

    if not result.success:
        raise ValueError(f"페이지 정보 추출 실패: {result.error}")

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    write_page_info_to_sheet(ws, result.page_meta)

    wb.save(filepath)
    print(f"페이지 정보 저장 완료: {filepath}")

    return filepath


# ============================================================
# 테스트
# ============================================================

if __name__ == "__main__":
    import sys
    import os
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

    from cursor import get_hwp_instance

    hwp = get_hwp_instance()
    if not hwp:
        print("[오류] 한글이 실행 중이지 않습니다.")
        exit(1)

    print("=== 페이지 정보 추출 테스트 ===\n")

    result = extract_page_info(hwp)

    if result.success:
        meta = result.page_meta
        print(f"용지 크기: {meta.page_size.width} x {meta.page_size.height} HWPUNIT")
        print(f"          {Unit.hwpunit_to_cm(meta.page_size.width):.1f} x {Unit.hwpunit_to_cm(meta.page_size.height):.1f} cm")
        print(f"용지 방향: {meta.page_size.orientation}")
        print(f"\n여백 (cm):")
        print(f"  왼쪽: {Unit.hwpunit_to_cm(meta.margin.left):.2f}")
        print(f"  오른쪽: {Unit.hwpunit_to_cm(meta.margin.right):.2f}")
        print(f"  위쪽: {Unit.hwpunit_to_cm(meta.margin.top):.2f}")
        print(f"  아래쪽: {Unit.hwpunit_to_cm(meta.margin.bottom):.2f}")
        print(f"\n본문 영역: {Unit.hwpunit_to_cm(meta.content_width):.1f} x {Unit.hwpunit_to_cm(meta.content_height):.1f} cm")

        # 엑셀 저장 테스트
        if HAS_OPENPYXL:
            import datetime
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            filepath = f"C:\\win32hwp\\test_page_{timestamp}.xlsx"
            save_page_info_to_excel(hwp, filepath)
    else:
        print(f"[오류] {result.error}")
