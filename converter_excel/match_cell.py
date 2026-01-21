# -*- coding: utf-8 -*-
"""한글 셀 스타일 → 엑셀 시트 저장 모듈

셀 배경색, 테두리, 여백, 글꼴 등의 스타일 정보를 엑셀 시트에 저장합니다.
"""

from dataclasses import dataclass, field
from typing import Optional, Dict, List, Tuple, Any

try:
    from openpyxl import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.worksheet.page import PageMargins
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from .page_meta import Unit


@dataclass
class CellStyleData:
    """셀 스타일 데이터"""
    list_id: int = 0

    # 위치 정보
    row: int = 0
    col: int = 0
    end_row: int = 0
    end_col: int = 0
    rowspan: int = 1
    colspan: int = 1

    # 물리 좌표 (HWPUNIT)
    x: int = 0
    y: int = 0
    width: int = 0
    height: int = 0

    # 배경색 (RGB)
    bg_color_rgb: Optional[Tuple[int, int, int]] = None

    # 셀 내부 여백 (HWPUNIT)
    margin_left: int = 0
    margin_right: int = 0
    margin_top: int = 0
    margin_bottom: int = 0

    # 글꼴 기본
    font_name: Optional[str] = None
    font_size_pt: float = 0
    font_bold: bool = False
    font_italic: bool = False
    font_color_rgb: Optional[Tuple[int, int, int]] = None

    # 글꼴 장식
    font_underline: int = 0        # 밑줄 타입 (0=없음, 1=실선, ...)
    font_underline_color: Optional[Tuple[int, int, int]] = None
    font_strikeout: int = 0        # 취소선 타입
    font_strikeout_color: Optional[Tuple[int, int, int]] = None
    font_outline: int = 0          # 외곽선 타입
    font_shadow: int = 0           # 그림자 타입
    font_emboss: bool = False      # 양각
    font_engrave: bool = False     # 음각
    font_superscript: bool = False # 위첨자
    font_subscript: bool = False   # 아래첨자

    # 자간/장평
    char_spacing: int = 0      # 자간 (%)
    char_ratio: int = 100      # 장평 (%)

    # 정렬
    align_horizontal: str = 'left'
    align_vertical: str = 'center'

    # 줄간격
    line_spacing: int = 0      # 줄간격 (%)

    # 테두리 (type, width)
    border_left_type: int = 0
    border_left_width: int = 0
    border_right_type: int = 0
    border_right_width: int = 0
    border_top_type: int = 0
    border_top_width: int = 0
    border_bottom_type: int = 0
    border_bottom_width: int = 0

    # 텍스트
    text: str = ""

    # 필드 이름 (배경색 없는 셀에만 설정)
    field_name: str = ""
    field_source: str = ""  # bookmark, A_B, A, B, random


@dataclass
class CellMatchResult:
    """셀 매칭 결과"""
    cells: List[CellStyleData] = field(default_factory=list)
    row_heights: List[int] = field(default_factory=list)  # HWPUNIT
    col_widths: List[int] = field(default_factory=list)   # HWPUNIT
    success: bool = False
    error: str = None


def extract_cell_style(hwp, list_id: int) -> CellStyleData:
    """한글 셀에서 스타일 정보 추출

    Args:
        hwp: HWP 객체
        list_id: 셀의 list_id

    Returns:
        CellStyleData 객체
    """
    style = CellStyleData(list_id=list_id)

    try:
        hwp.SetPos(list_id, 0, 0)

        # 1. 배경색 및 셀 내부 여백 (CellBorderFill)
        pset = hwp.HParameterSet.HCellBorderFill
        hwp.HAction.GetDefault("CellBorderFill", pset.HSet)

        # 배경색
        fill_attr = pset.FillAttr
        if fill_attr:
            bg_color = fill_attr.WinBrushFaceColor
            if bg_color and bg_color > 0 and bg_color != 4294967295:
                b = (bg_color >> 16) & 0xFF
                g = (bg_color >> 8) & 0xFF
                r = bg_color & 0xFF
                style.bg_color_rgb = (r, g, b)

            # 셀 내부 여백
            style.margin_left = getattr(fill_attr, 'InsideMarginLeft', 0) or 0
            style.margin_right = getattr(fill_attr, 'InsideMarginRight', 0) or 0
            style.margin_top = getattr(fill_attr, 'InsideMarginTop', 0) or 0
            style.margin_bottom = getattr(fill_attr, 'InsideMarginBottom', 0) or 0

        # 테두리
        style.border_left_type = pset.BorderTypeLeft
        style.border_left_width = pset.BorderWidthLeft
        style.border_right_type = pset.BorderTypeRight
        style.border_right_width = pset.BorderWidthRight
        style.border_top_type = pset.BorderTypeTop
        style.border_top_width = pset.BorderWidthTop
        style.border_bottom_type = pset.BorderTypeBottom
        style.border_bottom_width = pset.BorderWidthBottom

        # 2. 글자 속성 (HCharShape 파라미터셋)
        try:
            char_pset = hwp.HParameterSet.HCharShape
            hwp.HAction.GetDefault("CharShape", char_pset.HSet)

            # 기본 속성
            style.font_name = char_pset.FaceNameHangul
            style.font_size_pt = char_pset.Height / 100  # HWPUNIT -> pt
            style.font_bold = bool(char_pset.Bold)
            style.font_italic = bool(char_pset.Italic)

            # 글자색 (BGR -> RGB)
            text_color = char_pset.TextColor
            if text_color and text_color > 0:
                b = (text_color >> 16) & 0xFF
                g = (text_color >> 8) & 0xFF
                r = text_color & 0xFF
                style.font_color_rgb = (r, g, b)

            # 밑줄
            style.font_underline = char_pset.UnderlineType
            underline_color = char_pset.UnderlineColor
            if underline_color and underline_color > 0:
                b = (underline_color >> 16) & 0xFF
                g = (underline_color >> 8) & 0xFF
                r = underline_color & 0xFF
                style.font_underline_color = (r, g, b)

            # 취소선
            style.font_strikeout = char_pset.StrikeOutType
            strikeout_color = char_pset.StrikeOutColor
            if strikeout_color and strikeout_color > 0:
                b = (strikeout_color >> 16) & 0xFF
                g = (strikeout_color >> 8) & 0xFF
                r = strikeout_color & 0xFF
                style.font_strikeout_color = (r, g, b)

            # 외곽선, 그림자, 양각/음각
            style.font_outline = char_pset.OutLineType
            style.font_shadow = char_pset.ShadowType
            style.font_emboss = bool(char_pset.Emboss)
            style.font_engrave = bool(char_pset.Engrave)

            # 위첨자/아래첨자
            style.font_superscript = bool(char_pset.SuperScript)
            style.font_subscript = bool(char_pset.SubScript)

            # 자간/장평 (한글 기준)
            style.char_spacing = char_pset.SpacingHangul
            style.char_ratio = char_pset.RatioHangul
        except:
            pass

        # 3. 정렬/줄간격 (텍스트 선택 후 ParaShape)
        try:
            # 위치 재설정 후 텍스트 선택
            hwp.SetPos(list_id, 0, 0)
            hwp.HAction.Run("MoveSelParaEnd")
            ps = hwp.ParaShape
            if ps:
                align_val = ps.Item("AlignType")
                if align_val is not None:
                    # 0=양쪽, 1=왼쪽, 2=오른쪽, 3=가운데, 4=배분, 5=나눔
                    align_map = {0: 'justify', 1: 'left', 2: 'right', 3: 'center', 4: 'distribute', 5: 'divide'}
                    style.align_horizontal = align_map.get(align_val, 'left')

                line_sp = ps.Item("LineSpacing")
                if line_sp is not None:
                    style.line_spacing = line_sp
            hwp.HAction.Run("Cancel")
        except:
            pass

        # 4. 셀 세로 정렬 - CellShape.VertAlign이 항상 0 반환하는 문제 있음
        # 기본값 'center' 사용 (CellStyleData에서 설정됨)

    except Exception as e:
        pass

    return style


def get_cell_text(hwp, list_id: int) -> str:
    """셀 텍스트 추출 (SelectAll 사용)"""
    try:
        hwp.SetPos(list_id, 0, 0)

        # SelectAll로 셀 전체 선택
        hwp.HAction.Run("SelectAll")

        # 선택된 텍스트 가져오기 (saveblock 옵션 사용)
        text = hwp.GetTextFile("TEXT", "saveblock")

        # 선택 해제
        hwp.HAction.Run("Cancel")

        # 텍스트 정리
        if text:
            text = text.strip().replace('\r\n', ' ').replace('\r', ' ').replace('\n', ' ')
            return text
        return ""
    except:
        return ""


def write_cell_styles_to_sheet(ws: 'Worksheet', cells: List[CellStyleData],
                                row_heights: List[int] = None,
                                col_widths: List[int] = None):
    """셀 스타일 정보를 엑셀 시트에 기록

    Args:
        ws: openpyxl Worksheet 객체
        cells: CellStyleData 리스트
        row_heights: 행 높이 리스트 (HWPUNIT)
        col_widths: 열 너비 리스트 (HWPUNIT)
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl이 필요합니다: pip install openpyxl")

    # 스타일
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 헤더
    headers = [
        # 위치 정보
        "list_id", "row", "col", "end_row", "end_col", "rowspan", "colspan",
        # 물리 좌표
        "x", "y", "width", "height", "width_pt", "height_pt",
        # 배경색
        "bg_color",
        # 셀 내부 여백 (pt)
        "margin_L", "margin_R", "margin_T", "margin_B",
        # 글꼴 기본
        "font_name", "font_size", "bold", "italic", "font_color",
        # 글꼴 장식
        "underline", "strikeout", "outline", "shadow", "emboss", "engrave",
        "superscript", "subscript",
        # 자간/장평/줄간격
        "char_spacing", "char_ratio", "line_spacing",
        # 정렬
        "align_h", "align_v",
        # 테두리
        "border_L", "border_R", "border_T", "border_B",
        # 텍스트
        "text",
        # 필드
        "field_name", "field_source"
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # 셀 데이터
    sorted_cells = sorted(cells, key=lambda c: (c.row, c.col))

    for row_idx, cell_data in enumerate(sorted_cells, 2):
        values = [
            # 위치 정보
            cell_data.list_id,
            cell_data.row,
            cell_data.col,
            cell_data.end_row,
            cell_data.end_col,
            cell_data.rowspan,
            cell_data.colspan,
            # 물리 좌표
            cell_data.x,
            cell_data.y,
            cell_data.width,
            cell_data.height,
            round(cell_data.width / 100, 1),   # width_pt
            round(cell_data.height / 100, 1),  # height_pt
            # 배경색
            f"#{cell_data.bg_color_rgb[0]:02X}{cell_data.bg_color_rgb[1]:02X}{cell_data.bg_color_rgb[2]:02X}" if cell_data.bg_color_rgb else "",
            # 셀 내부 여백 (pt)
            round(cell_data.margin_left / 100, 1),
            round(cell_data.margin_right / 100, 1),
            round(cell_data.margin_top / 100, 1),
            round(cell_data.margin_bottom / 100, 1),
            # 글꼴 기본
            cell_data.font_name or "",
            cell_data.font_size_pt,
            "Y" if cell_data.font_bold else "",
            "Y" if cell_data.font_italic else "",
            f"#{cell_data.font_color_rgb[0]:02X}{cell_data.font_color_rgb[1]:02X}{cell_data.font_color_rgb[2]:02X}" if cell_data.font_color_rgb else "",
            # 글꼴 장식
            cell_data.font_underline if cell_data.font_underline else "",
            cell_data.font_strikeout if cell_data.font_strikeout else "",
            cell_data.font_outline if cell_data.font_outline else "",
            cell_data.font_shadow if cell_data.font_shadow else "",
            "Y" if cell_data.font_emboss else "",
            "Y" if cell_data.font_engrave else "",
            "Y" if cell_data.font_superscript else "",
            "Y" if cell_data.font_subscript else "",
            # 자간/장평/줄간격
            cell_data.char_spacing,
            cell_data.char_ratio,
            cell_data.line_spacing,
            # 정렬
            cell_data.align_horizontal,
            cell_data.align_vertical,
            # 테두리 (type/width)
            f"{cell_data.border_left_type}/{cell_data.border_left_width}",
            f"{cell_data.border_right_type}/{cell_data.border_right_width}",
            f"{cell_data.border_top_type}/{cell_data.border_top_width}",
            f"{cell_data.border_bottom_type}/{cell_data.border_bottom_width}",
            # 텍스트
            cell_data.text[:50] if cell_data.text else "",  # 50자 제한
            # 필드
            cell_data.field_name or "",
            cell_data.field_source or ""
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

            # 배경색이 있는 셀 표시
            if col_idx == 14 and cell_data.bg_color_rgb:  # bg_color 열
                r, g, b = cell_data.bg_color_rgb
                cell.fill = PatternFill(start_color=f"{r:02X}{g:02X}{b:02X}",
                                        end_color=f"{r:02X}{g:02X}{b:02X}",
                                        fill_type="solid")

    # 열 너비 조정
    col_widths_excel = [
        8, 5, 5, 7, 7, 7, 7,     # 위치 (7개)
        8, 8, 8, 8, 8, 8,        # 물리 좌표 (6개)
        10,                       # 배경색 (1개)
        7, 7, 7, 7,              # 여백 (4개)
        12, 8, 5, 5, 10,         # 글꼴 기본: name, size, bold, italic, color (5개)
        8, 8, 7, 7, 7, 7, 8, 8,  # 글꼴 장식: underline, strikeout, outline, shadow, emboss, engrave, superscript, subscript (8개)
        10, 10, 10,              # 자간/장평/줄간격 (3개)
        8, 8,                     # 정렬 (2개)
        8, 8, 8, 8,              # 테두리 (4개)
        25,                       # 텍스트 (1개)
        25, 10                    # 필드: field_name, field_source (2개)
    ]
    for col_idx, width in enumerate(col_widths_excel, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 인쇄 설정
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(
        left=0.3, right=0.3, top=0.3, bottom=0.3,
        header=0.2, footer=0.2
    )


def write_row_col_sizes_to_sheet(ws: 'Worksheet', row_heights: List[int], col_widths: List[int]):
    """행 높이와 열 너비를 별도 시트에 기록

    Args:
        ws: openpyxl Worksheet 객체
        row_heights: 행 높이 리스트 (HWPUNIT)
        col_widths: 열 너비 리스트 (HWPUNIT)
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl이 필요합니다: pip install openpyxl")

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 행 높이 섹션
    ws.cell(row=1, column=1, value="행 높이").font = header_font
    headers_row = ["행 번호", "HWPUNIT", "pt", "cm"]
    for col, header in enumerate(headers_row, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for idx, height in enumerate(row_heights):
        row = idx + 3
        ws.cell(row=row, column=1, value=idx + 1).border = thin_border
        ws.cell(row=row, column=2, value=height).border = thin_border
        ws.cell(row=row, column=3, value=round(height / 100, 2)).border = thin_border
        ws.cell(row=row, column=4, value=round(height / 2834.6, 2)).border = thin_border

    # 열 너비 섹션 (오른쪽에)
    start_col = 6
    ws.cell(row=1, column=start_col, value="열 너비").font = header_font
    headers_col = ["열 번호", "HWPUNIT", "pt", "cm", "문자(chars)"]
    for col, header in enumerate(headers_col, start_col):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for idx, width in enumerate(col_widths):
        row = idx + 3
        ws.cell(row=row, column=start_col, value=idx + 1).border = thin_border
        ws.cell(row=row, column=start_col + 1, value=width).border = thin_border
        ws.cell(row=row, column=start_col + 2, value=round(width / 100, 2)).border = thin_border
        ws.cell(row=row, column=start_col + 3, value=round(width / 2834.6, 2)).border = thin_border
        ws.cell(row=row, column=start_col + 4, value=round(width / 700, 2)).border = thin_border

    # 열 너비 조정
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 8
    ws.column_dimensions['I'].width = 8
    ws.column_dimensions['J'].width = 10

    # 인쇄 설정
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def apply_cell_style_to_excel_cell(ws: 'Worksheet', excel_row: int, excel_col: int,
                                    cell_data: CellStyleData):
    """셀 스타일을 엑셀 셀에 적용

    Args:
        ws: Worksheet 객체
        excel_row: 엑셀 행 번호 (1-based)
        excel_col: 엑셀 열 번호 (1-based)
        cell_data: CellStyleData 객체
    """
    if not HAS_OPENPYXL:
        return

    cell = ws.cell(row=excel_row, column=excel_col)

    # 배경색
    if cell_data.bg_color_rgb:
        r, g, b = cell_data.bg_color_rgb
        hex_color = f"{r:02X}{g:02X}{b:02X}"
        cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

    # 글꼴
    font_kwargs = {}
    if cell_data.font_name:
        font_kwargs['name'] = cell_data.font_name
    if cell_data.font_size_pt > 0:
        font_kwargs['size'] = cell_data.font_size_pt
    if cell_data.font_bold:
        font_kwargs['bold'] = True
    if cell_data.font_italic:
        font_kwargs['italic'] = True
    if cell_data.font_underline:
        font_kwargs['underline'] = 'single'
    if cell_data.font_color_rgb:
        r, g, b = cell_data.font_color_rgb
        font_kwargs['color'] = f"{r:02X}{g:02X}{b:02X}"

    if font_kwargs:
        cell.font = Font(**font_kwargs)

    # 정렬
    h_align_map = {'left': 'left', 'center': 'center', 'right': 'right',
                   'justify': 'justify', 'distribute': 'distributed'}
    v_align_map = {'top': 'top', 'center': 'center', 'bottom': 'bottom'}

    cell.alignment = Alignment(
        horizontal=h_align_map.get(cell_data.align_horizontal, 'left'),
        vertical=v_align_map.get(cell_data.align_vertical, 'center'),
        wrap_text=True
    )


# ============================================================
# 테스트
# ============================================================

if __name__ == "__main__":
    import sys
    import os
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

    from cursor import get_hwp_instance
    from table.table_info import TableInfo, MOVE_RIGHT_OF_CELL, MOVE_DOWN_OF_CELL

    hwp = get_hwp_instance()
    if not hwp:
        print("[오류] 한글이 실행 중이지 않습니다.")
        exit(1)

    table_info = TableInfo(hwp, debug=False)
    if not table_info.is_in_table():
        print("[오류] 커서가 표 안에 있지 않습니다.")
        exit(1)

    print("=== 셀 스타일 추출 테스트 ===\n")

    # 첫 셀로 이동
    table_info.move_to_first_cell()
    first_id = hwp.GetPos()[0]

    # BFS로 모든 셀 순회
    visited = set()
    queue = [first_id]
    visited.add(first_id)
    cells = []

    while queue and len(cells) < 50:
        cur_id = queue.pop(0)

        style = extract_cell_style(hwp, cur_id)
        style.text = get_cell_text(hwp, cur_id)
        cells.append(style)

        if style.bg_color_rgb:
            print(f"셀 {len(cells)}: bg={style.bg_color_rgb}, font={style.font_name}, size={style.font_size_pt}pt")

        # 다음 셀 탐색
        for move_cmd in [MOVE_RIGHT_OF_CELL, MOVE_DOWN_OF_CELL]:
            hwp.SetPos(cur_id, 0, 0)
            hwp.MovePos(move_cmd, 0, 0)
            next_id = hwp.GetPos()[0]
            if next_id != cur_id and next_id not in visited:
                visited.add(next_id)
                queue.append(next_id)

    print(f"\n총 {len(cells)}개 셀 탐색 완료")

    # 엑셀 저장 테스트
    if HAS_OPENPYXL:
        from openpyxl import Workbook
        import datetime

        wb = Workbook()
        ws = wb.active
        ws.title = "_cell_styles"

        write_cell_styles_to_sheet(ws, cells)

        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filepath = f"C:\\win32hwp\\test_cell_styles_{timestamp}.xlsx"
        wb.save(filepath)
        print(f"\n저장 완료: {filepath}")
